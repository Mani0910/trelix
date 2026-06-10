import os
import threading
import time

import openpyxl
import paramiko


EXCEL_SOURCE = r"\\10.211.34.118\api\trelix\mbg"
LOCAL_FILE = "fireeye.sh"
ROOT_PASSWORD = "Mitel5000"
REMOTE_PATH = "/root"
RUN_SCRIPT = True

HEADER_ALIASES = {
	"ip": ["server", "server ip", "ip", "host", "hostname"],
	"username": ["mbg username", "mivb username", "putty username", "username", "user"],
	"password": ["mbg password", "mivb password", "putty password", "password", "pass"],
}


def resolve_excel_file(source_path):
	"""Return the Excel file path from either a direct file path or a shared folder."""
	if os.path.isfile(source_path):
		return source_path

	if not os.path.isdir(source_path):
		print(f"[-] Excel path not found: {source_path}")
		return None

	candidates = []
	for name in os.listdir(source_path):
		lower_name = name.lower()
		if lower_name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
			candidates.append(os.path.join(source_path, name))

	if not candidates:
		print(f"[-] No Excel file found in: {source_path}")
		return None

	candidates.sort()
	return candidates[0]


def get_value(entry, keys):
	for key in keys:
		value = entry.get(key, "")
		if value:
			return value
	return ""


def load_servers(source_path):
	"""Read MBG server details from the shared Excel file."""
	excel_path = resolve_excel_file(source_path)
	if not excel_path:
		return []

	wb = openpyxl.load_workbook(excel_path)
	ws = wb.active

	header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
	if not header_row:
		print(f"[-] Excel file is empty: {excel_path}")
		return []

	headers = [str(cell).strip().lower() if cell else "" for cell in header_row]

	servers = []
	for row in ws.iter_rows(min_row=2, values_only=True):
		if not row or not row[0]:
			continue

		entry = dict(zip(headers, [str(value).strip() if value else "" for value in row]))
		ip = get_value(entry, HEADER_ALIASES["ip"])
		username = get_value(entry, HEADER_ALIASES["username"])
		password = get_value(entry, HEADER_ALIASES["password"])

		if not ip or not username or not password:
			print(f"[-] Skipping incomplete row for server: {ip or '<missing ip>'}")
			continue

		servers.append(
			{
				"ip": ip,
				"username": username,
				"password": password,
			}
		)

	print(f"Loaded {len(servers)} server(s) from: {excel_path}")
	return servers


def try_connect(ip, username, password):
	ssh = paramiko.SSHClient()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

	try:
		ssh.connect(
			hostname=ip,
			username=username,
			password=password,
			timeout=10,
			allow_agent=False,
			look_for_keys=False,
		)
		return ssh
	except paramiko.AuthenticationException:
		transport = paramiko.Transport((ip, 22))
		try:
			transport.start_client(timeout=10)
			transport.auth_interactive(
				username,
				lambda title, instructions, prompts: [password] * len(prompts),
			)
			if not transport.is_authenticated():
				raise paramiko.AuthenticationException("interactive authentication failed")

			ssh._transport = transport
			return ssh
		except Exception:
			transport.close()
			ssh.close()
			raise
	except Exception:
		ssh.close()
		raise


def wait_for_shell_output(shell, delay=1.0):
	time.sleep(delay)
	output = ""
	while shell.recv_ready():
		output += shell.recv(65535).decode(errors="replace")
		time.sleep(0.2)
	return output


def process_server(server):
	ip = server["ip"]
	username = server["username"]
	password = server["password"]

	try:
		ssh = try_connect(ip, username, password)
		print(f"{ip} - Login Successful")

		shell = ssh.invoke_shell()
		shell.settimeout(5)
		wait_for_shell_output(shell)

		shell.send("su -\n")
		output = wait_for_shell_output(shell)

		if "password" in output.lower():
			shell.send(ROOT_PASSWORD + "\n")
			output += wait_for_shell_output(shell, delay=2.0)

		if "#" not in output:
			print(f"{ip} - Root - Login Failed")
			ssh.close()
			return

		print(f"{ip} - Root - Login Successful")

		filename = os.path.basename(LOCAL_FILE)
		tmp_path = f"/tmp/{filename}"
		remote_path = f"{REMOTE_PATH.rstrip('/')}/{filename}"

		sftp = ssh.open_sftp()
		try:
			sftp.put(LOCAL_FILE, tmp_path)
		finally:
			sftp.close()

		setup_commands = [
			f"mv {tmp_path} {remote_path}",
			f"chmod 777 {remote_path}",
			f"ls -l {remote_path}",
		]

		for command in setup_commands:
			shell.send(command + "\n")
			command_output = wait_for_shell_output(shell)
			for line in command_output.splitlines():
				stripped = line.strip()
				if stripped:
					print(f"{ip} - {stripped}")

		if RUN_SCRIPT:
			shell.send(f"bash {remote_path}\n")
			start_time = time.time()
			timeout = 240

			while time.time() - start_time < timeout:
				if not shell.recv_ready():
					time.sleep(0.5)
					continue

				chunk = shell.recv(65535).decode(errors="replace")

				if "are you sure you want to continue connecting" in chunk.lower():
					shell.send("yes\n")
					continue

				if "password:" in chunk.lower() and "checking" not in chunk.lower():
					shell.send(ROOT_PASSWORD + "\n")
					continue

				for line in chunk.splitlines():
					stripped = line.strip()
					if not stripped:
						continue
					if stripped == "------------------------------":
						continue
					print(f"{ip} - {stripped}")

				if "Execution is completed" in chunk:
					break

		ssh.close()
	except Exception as exc:
		print(f"{ip} - Error: {exc}")


def sort_key(ip_address):
	try:
		return [int(part) for part in ip_address.split(".")]
	except ValueError:
		return [ip_address]


def main():
	if not os.path.exists(LOCAL_FILE):
		print(f"[-] Local script not found: {LOCAL_FILE}")
		return

	servers = load_servers(EXCEL_SOURCE)
	if not servers:
		print("[-] No servers loaded. Check the shared Excel file.")
		return

	servers.sort(key=lambda server: sort_key(server["ip"]))

	print(f"Starting Trellix check on {len(servers)} server(s)...\n")
	threads = []
	for server in servers:
		thread = threading.Thread(target=process_server, args=(server,))
		thread.start()
		threads.append(thread)

	for thread in threads:
		thread.join()

	print("\nAll servers processed.")


if __name__ == "__main__":
	main()
