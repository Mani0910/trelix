import paramiko
import time
import os
import threading
import openpyxl
from config import *

def load_servers(filepath):
    """Read server list from Excel file."""
    if not os.path.exists(filepath):
        print(f"[-] Excel file not found: {filepath}")
        return []

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    servers = []
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # skip empty rows
            continue
        entry = dict(zip(headers, [str(v).strip() if v else "" for v in row]))
        servers.append({
            "ip":             entry.get("server ip", ""),
            "admin_username": entry.get("putty username", ""),
            "admin_password": entry.get("putty password", ""),
            "root_password":  entry.get("mxone root password", ""),
        })

    return servers

def process_server(server):
    ip             = server["ip"]
    admin_username = server["admin_username"]
    admin_password = server["admin_password"]
    root_password  = server["root_password"]

    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        try:
            ssh.connect(
                hostname=ip,
                username=admin_username,
                password=admin_password,
                timeout=10,
                allow_agent=False,
                look_for_keys=False
            )
            print(f"{ip} - Login Successful")
        except paramiko.AuthenticationException:
            # Fallback: try keyboard-interactive auth
            try:
                transport = paramiko.Transport((ip, 22))
                transport.start_client()
                transport.auth_interactive(
                    admin_username,
                    lambda title, instructions, prompts: [admin_password] * len(prompts)
                )
                if not transport.is_authenticated():
                    print(f"{ip} - Login Failed (user: {admin_username})")
                    transport.close()
                    return
                # Wrap transport in SSHClient
                ssh._transport = transport
                print(f"{ip} - Login Successful")
            except Exception as e2:
                print(f"{ip} - Login Failed (user: {admin_username} | {e2})")
                return
        except Exception as e:
            print(f"{ip} - Login Failed ({e})")
            return

        # Open interactive shell
        shell = ssh.invoke_shell()
        time.sleep(1)
        shell.recv(9999)  # clear banner

        shell.send("su -\n")
        time.sleep(1)
        output = shell.recv(9999).decode()

        if "Password" in output or "password" in output:
            shell.send(root_password + "\n")
            time.sleep(2)
            output = shell.recv(9999).decode()

        if "#" not in output:
            print(f"{ip} - Root - Login Failed")
            ssh.close()
            return

        print(f"{ip} - Root - Login Successful")

        # Upload file silently
        filename = os.path.basename(LOCAL_FILE)
        tmp_path = f"/tmp/{filename}"
        remote_file_path = REMOTE_PATH.rstrip("/") + "/" + filename

        sftp = ssh.open_sftp()
        sftp.put(LOCAL_FILE, tmp_path)
        sftp.close()

        # Prepare and discard output of setup commands
        shell.settimeout(5)
        for cmd in [
            f"chmod 777 {tmp_path}",
            f"mv {tmp_path} {remote_file_path}",
            f"chmod 777 {remote_file_path}",
        ]:
            shell.send(cmd + "\n")
            time.sleep(1)
            try:
                shell.recv(9999)
            except Exception:
                pass

        if RUN_SCRIPT:
            shell.send(f"bash {remote_file_path}\n")
            time.sleep(2)

            start = time.time()
            timeout = 180

            while time.time() - start < timeout:
                time.sleep(0.5)
                if not shell.recv_ready():
                    continue

                chunk = shell.recv(65535).decode(errors="replace")

                if "Are you sure you want to continue connecting" in chunk:
                    shell.send("yes\n")
                    time.sleep(1)
                    continue
                elif "password:" in chunk.lower() and "Checking" not in chunk:
                    shell.send("Mitel5000\n")
                    time.sleep(1)
                    continue

                # Print meaningful lines in real-time
                for line in chunk.splitlines():
                    stripped = line.strip()
                    if not stripped:
                        continue
                    if stripped == "------------------------------":
                        continue
                    if ":~ #" in stripped or "bash /root/" in stripped:
                        continue
                    print(f"{ip} - {stripped}")

                if "Execution is completed" in chunk:
                    break

        ssh.close()

    except Exception as e:
        print(f"{ip} - Error: {e}")

def main():
    servers = load_servers(SERVERS_FILE)
    if not servers:
        print("[-] No servers loaded. Check your Excel file.")
        return

    # Sort by IP in ascending order
    servers.sort(key=lambda s: [int(x) for x in s["ip"].split(".")])

    print(f"Starting Trellix check on {len(servers)} server(s)...\n")
    threads = []
    for server in servers:
        t = threading.Thread(target=process_server, args=(server,))
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    print("\nAll servers processed.")

if __name__ == "__main__":
    main()