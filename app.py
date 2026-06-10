from flask import Flask, request, Response, jsonify, send_file
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import paramiko
import time
import os
import sys
import io
import base64
import json
import csv
import logging
import secrets
import socket
import subprocess
import platform
import threading
import tempfile
import concurrent.futures
import queue
import openpyxl
import smtplib
from functools import wraps
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from werkzeug.utils import secure_filename
from smb.SMBConnection import SMBConnection
import winrm
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except ImportError:
    psycopg2 = None
    RealDictCursor = None


def get_base_path():
    """Get base path for bundled files (PyInstaller compatible)."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))


BASE_PATH = get_base_path()
APP_DIR = os.path.dirname(os.path.abspath(sys.executable)) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(APP_DIR, "data")
AUX_DIR = os.path.join(APP_DIR, "uploads", "audits")
PROFILE_TEMPLATES_FILE = os.path.join(DATA_DIR, "deployment_profile_templates.json")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(AUX_DIR, exist_ok=True)


def _get_index_file_path():
    """Pick a valid index.html and avoid stale/invalid files that cause blank UI."""
    candidates = [
        os.getenv("TRELLIX_INDEX_PATH", "").strip(),
        os.path.join(APP_DIR, 'index.html'),
        os.path.join(os.getcwd(), 'index.html'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'index.html'),
        os.path.join(BASE_PATH, 'index.html'),
    ]

    def _is_valid_index(path):
        if not path or not os.path.isfile(path):
            return False
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as fh:
                # Read a bounded amount; enough to validate expected app markers.
                text = fh.read(200000)
        except Exception:
            return False

        required_markers = (
            'id="landingPage"',
            'id="platformPage"',
            'data-app-ready',
        )
        return all(marker in text for marker in required_markers)

    for path in candidates:
        if _is_valid_index(path):
            return path

    # Last-resort fallback: return bundled path even if validation failed.
    return os.path.join(BASE_PATH, 'index.html')


def _serve_index():
    return send_file(_get_index_file_path())

app = Flask(__name__, static_folder=BASE_PATH, template_folder=BASE_PATH)
app.config['UPLOAD_FOLDER'] = os.path.join(APP_DIR, 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

logger = logging.getLogger("trelix")
if not logger.handlers:
    logging.basicConfig(level=logging.INFO, format='%(message)s')


def _mask_secret(value, keep=2):
    raw = str(value or "")
    if not raw:
        return ""
    if len(raw) <= keep:
        return "*" * len(raw)
    return ("*" * (len(raw) - keep)) + raw[-keep:]


def log_structured(event_type, **fields):
    payload = {
        "ts": datetime.utcnow().isoformat() + "Z",
        "event": event_type,
    }
    payload.update(fields)
    try:
        logger.info(json.dumps(payload, default=str))
    except Exception:
        logger.info(str(payload))

# Rate limiting: 100 requests per user (by IP)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["100 per minute"],
    storage_uri="memory://",
)

@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'

    # Ensure latest UI is always served after deployments.
    normalized_path = request.path.rstrip('/') or '/'
    if normalized_path in ('/', '/trelix', '/mivbtrelix', '/mbgtrelix', '/mxonetrelix', '/miv5000trelix', '/windowstrelix'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# Global queue for SSE events
event_queues = {}

PLATFORM_CONFIG = {
    "mxone": {
        "title": "MxOne Trellix Manager",
        "share_dir": os.getenv("SHARE_DIR_MXONE", "/upgrade_credentials/mxone"),
        "headers": {
            "ip": ["ip"],
            "username": ["putty username", "puttyusername"],
            "password": ["putty password", "puttypassword"],
            "root_password": ["root password", "rootpassword", "mxone root password"],
        },
        "remote_path": "/root",
        "script_password": "Mitel5000",
    },
    "mivb": {
        "title": "MiVB Trellix Manager",
        "share_dir": os.getenv("SHARE_DIR_MIVB", "/trelix_credentails/mivb"),
        "headers": {
            "ip": ["server", "server ip", "ip", "host", "hostname"],
            "username": ["mivb username", "putty username", "username", "user"],
            "password": ["mivb password", "putty password", "password", "pass"],
        },
        "remote_path": "/root",
        "root_password": "Mitel5000",
        "script_password": "Mitel5000",
    },
    "mbg": {
        "title": "MBG Trellix Manager",
        "share_dir": os.getenv("SHARE_DIR_MBG", "/trelix_credentails/mbg"),
        "headers": {
            "ip": ["server", "server ip", "ip", "host", "hostname"],
            "username": ["mbg username", "mivb username", "putty username", "username", "user"],
            "password": ["mbg password", "mivb password", "putty password", "password", "pass"],
        },
        "remote_path": "/root",
        "root_password": "Mitel5000",
        "script_password": "Mitel5000",
    },
    "miv5000": {
        "title": "MiV5000 Trellix Manager",
        "share_dir": os.getenv("SHARE_DIR_MIV5000", "/upgrade_credentials/miv5000"),
        "headers": {
            "ip": ["server ip", "ip", "server", "host", "hostname"],
            "username": ["putty username", "username", "user"],
            "password": ["putty password", "password", "pass"],
        },
        "remote_path": "/root",
        "root_password": "Mitel5000",
        "script_password": "Mitel5000",
    },
    "windows": {
        "title": "Windows Trellix Manager",
        "share_dir": os.getenv("SHARE_DIR_WINDOWS", "/trelix_credentials/windows"),
        "headers": {
            "ip": ["server ip", "ip", "server", "host", "hostname"],
            "username": ["windows username", "username", "user", "admin username"],
            "password": ["windows password", "password", "pass", "admin password"],
        },
        "winrm_port": int(os.getenv("WINRM_PORT", "5985")),
        "winrm_transport": os.getenv("WINRM_TRANSPORT", "ntlm"),
        "installer_unc": os.getenv("WINDOWS_INSTALLER_UNC", ""),
        "scp_enabled": str(os.getenv("WINDOWS_SCP_ENABLED", "0")).strip().lower() in ("1", "true", "yes", "on"),
        "scp_source_host": os.getenv("WINDOWS_SCP_SOURCE_HOST", "10.211.27.74").strip(),
        "scp_source_port": int(os.getenv("WINDOWS_SCP_SOURCE_PORT", "22")),
        "scp_source_user": os.getenv("WINDOWS_SCP_SOURCE_USER", "root").strip(),
        "scp_source_pass": os.getenv("WINDOWS_SCP_SOURCE_PASS", "").strip(),
        "scp_source_path": os.getenv("WINDOWS_SCP_SOURCE_PATH", "/home/Fireeye/IMAGE_HX_AGENT_WIN_36.30.37.zip").strip(),
        "package_basename": os.getenv("WINDOWS_PACKAGE_BASENAME", "IMAGE_HX_AGENT_WIN_36.30.37").strip(),
        "scp_target_dir": os.getenv("WINDOWS_SCP_TARGET_DIR", r"C:\Users\Administrator\Downloads").strip(),
        "scp_auto_setup_sshd": str(os.getenv("WINDOWS_SCP_AUTO_SETUP_SSHD", "1")).strip().lower() in ("1", "true", "yes", "on"),
        "scp_unc_fallback": str(os.getenv("WINDOWS_SCP_UNC_FALLBACK", "0")).strip().lower() in ("1", "true", "yes", "on"),
    },
}

# Email Configuration — parsed from config.py (key=value format)
def _load_config(path=None):
    if path is None:
        path = os.path.join(APP_DIR, "config.py")
    cfg = {}
    try:
        with open(path) as f:
            for line in f:
                line = line.strip()
                if "=" in line and not line.startswith("#"):
                    k, _, v = line.partition("=")
                    cfg[k.strip()] = v.strip()
    except Exception:
        pass
    return cfg

_cfg = _load_config()
EMAIL_CONFIG = {
    "smtp_server": _cfg.get('SMTP_SERVER', 'smtp.mitel.com'),
    "smtp_port": int(_cfg.get('SMTP_PORT', 587)),
    "sender": _cfg.get('SENDER_EMAIL', ''),
    "password": _cfg.get('SENDER_PASSWORD', ''),
    "recipient": _cfg.get('RECIPIENTS', ''),
}

SECURE_EXPORT_DIR = os.getenv("SECURE_EXPORT_DIR", "/secure_data")
SECURE_EXPORT_XLSX = os.path.join(SECURE_EXPORT_DIR, "trelix_data_secure.xlsx")
ENABLE_RESULT_EXPORTS = str(os.getenv("ENABLE_RESULT_EXPORTS", "0")).strip().lower() in ("1", "true", "yes", "on")

try:
    os.makedirs(SECURE_EXPORT_DIR, exist_ok=True)
except Exception:
    # Export path may be unavailable in some local runs; runtime writes are best-effort.
    pass

DB_HOST = os.getenv("DB_HOST", "")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_NAME = os.getenv("DB_NAME", "")
DB_USER = os.getenv("DB_USER", "")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")


def _db_enabled():
    return bool(psycopg2 and DB_HOST and DB_NAME and DB_USER and DB_PASSWORD)


def _get_db_conn():
    if not _db_enabled():
        return None
    return psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
    )


def _extract_user_identifier(req):
    user_name = (req.headers.get("X-User") or req.headers.get("X-Username") or "").strip()
    if not user_name:
        user_name = (req.remote_addr or "anonymous").strip()
    return user_name[:120]


def _set_session_result(session_id, ip, **fields):
    current = session_results.setdefault(session_id, {}).get(ip, {})
    current.update(fields)
    session_results.setdefault(session_id, {})[ip] = current


def _is_system_reachable(ip, port=22, timeout=5):
    """Check if system is reachable via a specific TCP port."""
    try:
        probe = socket.create_connection((ip, int(port)), timeout=timeout)
        probe.close()
        return True
    except Exception:
        return False


def _is_ip_pingable(ip, timeout=2):
    """Best-effort ICMP ping check for host reachability."""
    try:
        if platform.system().lower().startswith("win"):
            # Windows ping: timeout is per-reply in milliseconds.
            cmd = ["ping", "-n", "1", "-w", str(int(timeout * 1000)), ip]
        else:
            # Linux/macOS ping: timeout in seconds.
            cmd = ["ping", "-c", "1", "-W", str(int(timeout)), ip]
        completed = subprocess.run(cmd, capture_output=True, text=True, timeout=max(3, int(timeout) + 2))
        return completed.returncode == 0
    except Exception:
        return False


def _is_credentials_valid_result(info):
    """Treat credentials as valid only when authentication/root login clearly succeeded."""
    explicit = (info or {}).get("credential_valid", None)
    if explicit is True:
        return True
    if explicit is False:
        return False

    status = str((info or {}).get("status", "")).lower()
    message = str((info or {}).get("message", "")).lower()

    auth_fail_markers = [
        "authentication failed",
        "login failed",
        "root login failed",
        "invalid credential",
        "winrm session failed",
        "access denied",
        "permission denied",
    ]
    if any(m in message for m in auth_fail_markers):
        return False
    if status == "error":
        return False
    return True


def _credential_valid_label(info):
    """Return credential validity label for storage/reporting."""
    explicit = (info or {}).get("credential_valid", None)
    if explicit is True:
        return "valid"
    if explicit is False:
        return "invalid"

    status = str((info or {}).get("status", "")).strip().lower()
    message = str((info or {}).get("message", "")).strip().lower()

    auth_fail_markers = [
        "authentication failed",
        "login failed",
        "root login failed",
        "invalid credential",
        "access denied",
        "permission denied",
        "credentials invalid",
    ]
    if any(m in message for m in auth_fail_markers):
        return "invalid"

    # Deployment statuses usually don't carry explicit credential_valid.
    # If we reached installed/warning after authenticated actions, treat as valid.
    if status in ("installed", "warning"):
        return "valid"

    # Keep non-auth errors as unknown (e.g., timeout/unreachable/runtime issues).
    return "unknown"


def _precheck_details_to_results(precheck_details):
    """Convert precheck detail payload to result map used by storage/export."""
    result = {}
    for item in precheck_details or []:
        if not isinstance(item, dict):
            continue
        ip = str(item.get("ip") or "").strip()
        if not ip:
            continue
        state = str(item.get("state") or "").strip().lower()
        if state == "installed":
            status = "installed"
        elif state == "pending":
            status = "warning"
        else:
            status = "error"

        result[ip] = {
            "status": status,
            "message": str(item.get("reason") or "Precheck completed"),
            "version": str(item.get("version") or ""),
            "credential_valid": item.get("credential_valid", None),
        }
    return result


def _dedupe_servers_by_ip(servers):
    """Keep last row per IP (latest credentials) and avoid duplicate processing/saves."""
    by_ip = {}
    for server in servers or []:
        ip_raw = str((server or {}).get("ip") or "").strip()
        parts = ip_raw.split(".")
        if len(parts) != 4:
            continue
        try:
            nums = [int(p) for p in parts]
        except Exception:
            continue
        if not all(0 <= n <= 255 for n in nums):
            continue
        canonical_ip = ".".join(str(n) for n in nums)
        row = dict(server)
        row["ip"] = canonical_ip
        by_ip[canonical_ip] = row

    deduped = list(by_ip.values())
    deduped.sort(key=lambda s: [int(x) for x in s["ip"].split(".")])
    return deduped


def _create_deployment_run(req, platform, source_file):
    """Create app_users + deployment_runs records and return run_id (or None if DB unavailable)."""
    if not _db_enabled():
        return None

    conn = None
    try:
        conn = _get_db_conn()
        with conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                username = _extract_user_identifier(req)
                cur.execute(
                    """
                    INSERT INTO app_users (username, email)
                    VALUES (%s, %s)
                    RETURNING id
                    """,
                    (username, None),
                )
                user_id = cur.fetchone()["id"]

                cur.execute(
                    """
                    INSERT INTO deployment_runs (user_id, vm_type, source_file, started_at)
                    VALUES (%s, %s, %s, NOW())
                    RETURNING id
                    """,
                    (user_id, platform, source_file),
                )
                run_id = cur.fetchone()["id"]
                return run_id
    except Exception as exc:
        print(f"[DB] create run failed: {exc}")
        return None
    finally:
        if conn:
            conn.close()


def _upsert_server_inventory(cur, server, platform, created_by_user_id=None, update_credentials=True, credential_valid=None):
    """Upsert server into inventory and return server_id.

    - Always stores the record (first time: valid or not).
    - If update_credentials=True and server exists, rotates current password
      to previous_* columns before saving the new credentials.
    """
    ip = server.get("ip")
    admin_username = server.get("admin_username", "")
    admin_password = server.get("admin_password", "")
    root_password = server.get("root_password", "")

    # Enforce one logical inventory record per IP to avoid duplicates.
    cur.execute(
        "SELECT id, putty_password_encrypted, root_password_encrypted FROM server_inventory WHERE ip_address = %s LIMIT 1",
        (ip,)
    )
    found = cur.fetchone()
    if found:
        if isinstance(found, dict):
            existing_id = found["id"]
            old_pass = found.get("putty_password_encrypted") or ""
            old_root = found.get("root_password_encrypted") or ""
        else:
            existing_id, old_pass, old_root = found[0], found[1] or "", found[2] or ""

        if update_credentials:
            # Rotate current → previous only when the password actually changed.
            new_prev_pass = old_pass if old_pass != admin_password else None
            new_prev_root = old_root if old_root != root_password else None
            cur.execute(
                """
                UPDATE server_inventory
                SET vm_type = %s,
                    putty_username = %s,
                    putty_password_encrypted = %s,
                    previous_putty_password = COALESCE(%s, previous_putty_password),
                    root_password_encrypted = %s,
                    previous_root_password = COALESCE(%s, previous_root_password),
                    credential_valid = %s,
                    created_by = COALESCE(%s, created_by),
                    last_credential_update = NOW()
                WHERE id = %s
                """,
                (platform, admin_username, admin_password, new_prev_pass,
                 root_password, new_prev_root, credential_valid,
                 created_by_user_id, existing_id),
            )
        else:
            # Not valid — still update validity flag and vm_type, keep passwords as-is.
            cur.execute(
                """
                UPDATE server_inventory
                SET vm_type = %s,
                    credential_valid = %s
                WHERE id = %s
                """,
                (platform, credential_valid, existing_id),
            )
        return existing_id

    # First-time insert — store regardless of credential validity.
    cur.execute(
        """
        INSERT INTO server_inventory (
            ip_address,
            vm_type,
            putty_username,
            putty_password_encrypted,
            root_password_encrypted,
            credential_valid,
            created_by,
            created_at,
            last_credential_update
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        RETURNING id
        """,
        (ip, platform, admin_username, admin_password, root_password,
         credential_valid, created_by_user_id),
    )
    row = cur.fetchone()
    return row["id"] if isinstance(row, dict) else row[0]


def _export_secure_xlsx(platform, servers, per_ip_results):
    """Upsert server data to secure XLSX with unique IP rows.

    Rules:
    - Always stores every server (first time: valid or not).
    - Rotates putty_password → previous_putty_password when a new valid password is used.
    - Color-codes the credentials_valid cell: green=valid, red=invalid, orange=unknown.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
    if not servers:
        return

    try:
        os.makedirs(SECURE_EXPORT_DIR, exist_ok=True)
    except Exception:
        return

    headers = [
        "ip",
        "vm_type",
        "putty_username",
        "putty_password",
        "previous_putty_password",
        "root_password",
        "previous_root_password",
        "credentials_valid",
        "trelix_installed",
        "trelix_version",
        "status",
        "message",
        "updated_at",
    ]

    # Color fills for credential validity
    fill_valid   = PatternFill("solid", fgColor="C6EFCE")  # green
    fill_invalid = PatternFill("solid", fgColor="FFC7CE")  # red
    fill_unknown = PatternFill("solid", fgColor="FFEB9C")  # orange/yellow
    font_valid   = Font(color="276221", bold=True)
    font_invalid = Font(color="9C0006", bold=True)
    font_unknown = Font(color="9C6500", bold=True)
    fill_status_ok = PatternFill("solid", fgColor="D9EAD3")
    fill_status_warn = PatternFill("solid", fgColor="FCE5CD")
    fill_status_err = PatternFill("solid", fgColor="F4CCCC")
    fill_normal = PatternFill(fill_type=None)
    font_normal = Font(color="000000", bold=False)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)

    def _looks_like_ipv4(value):
        s = str(value or "").strip()
        parts = s.split('.')
        if len(parts) != 4:
            return False
        try:
            nums = [int(p) for p in parts]
        except Exception:
            return False
        return all(0 <= n <= 255 for n in nums)

    if os.path.exists(SECURE_EXPORT_XLSX):
        wb = openpyxl.load_workbook(SECURE_EXPORT_XLSX)
        if "trelix_data" in wb.sheetnames:
            ws = wb["trelix_data"]
        else:
            ws = wb.active
            ws.title = "trelix_data"
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "trelix_data"

    # Keep exactly one sheet in the workbook.
    for sheet_name in list(wb.sheetnames):
        if sheet_name != "trelix_data":
            del wb[sheet_name]
    ws = wb["trelix_data"]

    # Auto-repair malformed workbook patterns (e.g., shifted headers / millions of duplicate rows).
    header_probe = [str(ws.cell(row=1, column=i).value or "").strip().lower() for i in range(1, min(ws.max_column, 20) + 1)]
    has_shifted_headers = (ws.cell(row=1, column=1).value in (None, "") and len(header_probe) > 1 and header_probe[1] == "ip")
    suspicious_size = ws.max_row >= 500000
    if has_shifted_headers or suspicious_size:
        repaired = wb.create_sheet("trelix_data_repaired")
        repaired.append(headers)
        for c in range(1, len(headers) + 1):
            cell = repaired.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        old_headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v:
                old_headers[str(v).strip().lower()] = c

        def old_col(name):
            return old_headers.get(name.lower())

        seen = set()
        scan_limit = min(ws.max_row, 50000)
        for r in range(2, scan_limit + 1):
            ic = old_col("ip")
            if not ic:
                break
            ip_val = ws.cell(row=r, column=ic).value
            if not _looks_like_ipv4(ip_val):
                continue
            ip = str(ip_val).strip()
            if ip in seen:
                continue
            seen.add(ip)

            row_values = [
                ip,
                platform,
                ws.cell(row=r, column=old_col("putty_username") or 1).value or "",
                ws.cell(row=r, column=old_col("putty_password") or 1).value or "",
                ws.cell(row=r, column=old_col("previous_putty_password") or 1).value or "",
                ws.cell(row=r, column=old_col("root_password") or 1).value or "",
                ws.cell(row=r, column=old_col("previous_root_password") or 1).value or "",
                ws.cell(row=r, column=old_col("credentials_valid") or 1).value or "UNKNOWN",
                ws.cell(row=r, column=old_col("trelix_installed") or 1).value or "no",
                ws.cell(row=r, column=old_col("trelix_version") or 1).value or "",
                ws.cell(row=r, column=old_col("status") or 1).value or "",
                ws.cell(row=r, column=old_col("message") or 1).value or "",
                ws.cell(row=r, column=old_col("updated_at") or 1).value or "",
            ]
            repaired.append(row_values)

        wb.remove(ws)
        repaired.title = "trelix_data"
        ws = repaired

    if ws.max_row == 0:
        ws.append(headers)
    if ws.max_row >= 1 and str(ws.cell(row=1, column=1).value or "").strip().lower() != "ip":
        # Ensure canonical header row shape.
        ws.delete_rows(1, 1)
        ws.insert_rows(1, 1)
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=h)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Ensure all expected headers exist (handles older files).
    existing_headers = {}
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col_idx).value
        if raw:
            existing_headers[str(raw).strip().lower()] = col_idx

    for h in headers:
        key = h.lower()
        if key not in existing_headers:
            nc = ws.max_column + 1
            cell = ws.cell(row=1, column=nc, value=h)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            existing_headers[key] = nc

    def hcol(name):
        return existing_headers[name.lower()]

    # Remove accidental header-like data rows from older malformed exports.
    for row_idx in range(ws.max_row, 1, -1):
        ip_val = str(ws.cell(row=row_idx, column=hcol("ip")).value or "").strip().lower()
        vm_val = str(ws.cell(row=row_idx, column=hcol("vm_type")).value or "").strip().lower()
        user_val = str(ws.cell(row=row_idx, column=hcol("putty_username")).value or "").strip().lower()
        if ip_val == "ip" and vm_val == "vm_type" and user_val == "putty_username":
            ws.delete_rows(row_idx, 1)

    # Remove duplicate IP rows already present in the workbook.
    seen_ips = set()
    for row_idx in range(ws.max_row, 1, -1):
        key = ws.cell(row=row_idx, column=hcol("ip")).value
        if not key or not _looks_like_ipv4(key):
            continue
        ip_key = ".".join(str(int(p)) for p in str(key).strip().split("."))
        if ip_key in seen_ips:
            ws.delete_rows(row_idx, 1)
            continue
        seen_ips.add(ip_key)
        if str(key).strip() != ip_key:
            ws.cell(row=row_idx, column=hcol("ip"), value=ip_key)

    # Build IP → row index map from existing data.
    ip_row = {}
    for row_idx in range(2, ws.max_row + 1):
        key = ws.cell(row=row_idx, column=hcol("ip")).value
        if key and _looks_like_ipv4(key):
            ip_row[".".join(str(int(p)) for p in str(key).strip().split("."))] = row_idx

    for server in _dedupe_servers_by_ip(servers):
        ip = server.get("ip")
        info = per_ip_results.get(ip, {})

        status    = str(info.get("status", ""))
        version   = str(info.get("version", "") or "")
        message   = str(info.get("message", ""))
        installed = "yes" if status == "installed" else "no"
        validity  = _credential_valid_label(info)  # "valid" / "invalid" / "unknown"

        new_pass  = server.get("admin_password", "")
        new_root  = server.get("root_password", "")

        # Platform policy: mivb/mbg/miv5000 always use fixed script/root password.
        if platform in ("mivb", "mbg", "miv5000"):
            new_root = "Mitel5000"

        # Determine previous/current password handling.
        # Existing row: update current password only when credentials are VALID.
        # First row: keep provided password regardless, and mark VALID/INVALID/UNKNOWN.
        prev_pass = ""
        prev_root = ""
        current_pass = new_pass
        current_root = new_root
        if ip in ip_row:
            r = ip_row[ip]
            existing_pass = ws.cell(row=r, column=hcol("putty_password")).value or ""
            existing_root = ws.cell(row=r, column=hcol("root_password")).value or ""
            existing_prev_pass = ws.cell(row=r, column=hcol("previous_putty_password")).value or ""
            existing_prev_root = ws.cell(row=r, column=hcol("previous_root_password")).value or ""

            # Only rotate to previous when credentials are valid AND changed.
            if validity == "valid":
                if existing_pass and existing_pass != new_pass:
                    prev_pass = existing_pass
                else:
                    prev_pass = existing_prev_pass
                if existing_root and existing_root != new_root:
                    prev_root = existing_root
                else:
                    prev_root = existing_prev_root
            else:
                # Invalid/unknown: preserve existing passwords and previous history.
                current_pass = existing_pass
                current_root = existing_root
                prev_pass = existing_prev_pass
                prev_root = existing_prev_root

            # Enforce fixed root password policy in sheet for mivb/mbg/miv5000.
            if platform in ("mivb", "mbg", "miv5000"):
                current_root = "Mitel5000"

        values = {
            "ip":                      ip,
            "vm_type":                 platform,
            "putty_username":          server.get("admin_username", ""),
            "putty_password":          current_pass,
            "previous_putty_password": prev_pass,
            "root_password":           current_root,
            "previous_root_password":  prev_root,
            "credentials_valid":       validity.upper(),
            "trelix_installed":        installed,
            "trelix_version":          version,
            "status":                  status,
            "message":                 message,
            "updated_at":              datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        r = ip_row[ip] if ip in ip_row else ws.max_row + 1
        for col_name, val in values.items():
            cell = ws.cell(row=r, column=hcol(col_name), value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = fill_normal
            cell.font = font_normal

        # Apply color to credentials_valid cell
        valid_cell = ws.cell(row=r, column=hcol("credentials_valid"))
        valid_cell.alignment = Alignment(horizontal="center", vertical="center")
        if validity == "valid":
            valid_cell.fill = fill_valid
            valid_cell.font = font_valid
        elif validity == "invalid":
            valid_cell.fill = fill_invalid
            valid_cell.font = font_invalid
        else:
            valid_cell.fill = fill_unknown
            valid_cell.font = font_unknown

        status_cell = ws.cell(row=r, column=hcol("status"))
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_text = str(status).lower()
        if status_text == "installed":
            status_cell.fill = fill_status_ok
        elif status_text == "warning":
            status_cell.fill = fill_status_warn
        elif status_text == "error":
            status_cell.fill = fill_status_err

        if ip not in ip_row:
            ip_row[ip] = r

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"

    # Auto-fit column widths (bounded scan for performance on large sheets).
    width_scan_limit = min(ws.max_row, 500)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, width_scan_limit + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None:
                continue
            try:
                max_len = max(max_len, len(str(cell_val)))
            except Exception:
                pass
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    # Keep XLSX as protected important data (sheet + workbook structure password).
    admin_xlsx_password = os.getenv("ADMIN_XLSX_PASSWORD", "Mitel@123")

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=True)

    ws.protection.sheet = True
    ws.protection.set_password(admin_xlsx_password)
    ws.protection.enable()
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.insertHyperlinks = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = True
    ws.protection.pivotTables = False

    if getattr(wb, "security", None) is None:
        from openpyxl.workbook.protection import WorkbookProtection
        wb.security = WorkbookProtection()
    wb.security.lockStructure = True
    wb.security.workbookPassword = admin_xlsx_password

    # Atomic write: save to .tmp first, then replace the target file.
    # This prevents partial writes and works even when Excel has the file open
    # (Windows user-space locks are not enforced for Linux processes in Docker).
    import shutil
    tmp_path = SECURE_EXPORT_XLSX + ".tmp"
    try:
        wb.save(tmp_path)
    except Exception as exc:
        print(f"[XLSX] failed to write temp file {tmp_path}: {exc}")
        raise
    try:
        os.replace(tmp_path, SECURE_EXPORT_XLSX)
        print(f"[XLSX] saved: {SECURE_EXPORT_XLSX}")
    except Exception as exc:
        print(f"[XLSX] os.replace failed ({exc}), trying shutil.copy2")
        try:
            shutil.copy2(tmp_path, SECURE_EXPORT_XLSX)
            os.remove(tmp_path)
            print(f"[XLSX] saved via copy: {SECURE_EXPORT_XLSX}")
        except Exception as exc2:
            print(f"[XLSX] copy also failed: {exc2}; keeping temp export at {tmp_path}")
            # Do not fail deployment flow due to an external workbook lock.
            return


def _ensure_db_schema_runtime():
    """Best-effort runtime schema safety for existing DBs."""
    if not _db_enabled():
        return

    conn = None
    try:
        conn = _get_db_conn()
        with conn:
            with conn.cursor() as cur:
                cur.execute("ALTER TABLE deployment_results ADD COLUMN IF NOT EXISTS credential_valid BOOLEAN")
                cur.execute("ALTER TABLE server_inventory ADD COLUMN IF NOT EXISTS last_credential_update TIMESTAMPTZ DEFAULT NOW()")
                cur.execute("ALTER TABLE server_inventory ADD COLUMN IF NOT EXISTS previous_putty_password TEXT")
                cur.execute("ALTER TABLE server_inventory ADD COLUMN IF NOT EXISTS previous_root_password TEXT")
                cur.execute("ALTER TABLE server_inventory ADD COLUMN IF NOT EXISTS credential_valid BOOLEAN")
                # Keep latest result row per server to support unique upsert and avoid duplicates.
                cur.execute(
                    """
                    DELETE FROM deployment_results d
                    USING deployment_results d2
                    WHERE d.server_id = d2.server_id
                      AND d.id < d2.id
                    """
                )
                cur.execute(
                    """
                    CREATE UNIQUE INDEX IF NOT EXISTS uq_deployment_results_server_id
                    ON deployment_results (server_id)
                    """
                )
    except Exception as exc:
        print(f"[DB] runtime schema ensure failed: {exc}")
    finally:
        if conn:
            conn.close()


def _save_run_results_to_db(run_id, platform, servers, per_ip_results):
    """Persist server input and deployment outputs for a run."""
    if not servers:
        return

    conn = None
    try:
        if run_id and _db_enabled():
            _ensure_db_schema_runtime()
            conn = _get_db_conn()
            with conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT user_id FROM deployment_runs WHERE id = %s", (run_id,))
                    row = cur.fetchone() or {}
                    user_id = row.get("user_id") if isinstance(row, dict) else (row[0] if row else None)

                    for server in _dedupe_servers_by_ip(servers):
                        ip = server.get("ip")
                        if not ip:
                            continue
                        info = per_ip_results.get(ip, {})
                        credentials_ok = _is_credentials_valid_result(info)
                        credential_valid = info.get("credential_valid", None)
                        if credential_valid is None:
                            credential_valid = credentials_ok

                        server_id = _upsert_server_inventory(
                            cur,
                            server,
                            platform,
                            user_id,
                            update_credentials=credentials_ok,
                            credential_valid=credential_valid,
                        )

                        status = str(info.get("status", "warning"))
                        message = str(info.get("message", "No result captured"))
                        version = str(info.get("version", "") or "")
                        installed = (status == "installed")

                        cur.execute(
                            """
                            INSERT INTO deployment_results (
                                run_id,
                                server_id,
                                trelix_installed,
                                trelix_version,
                                status,
                                message,
                                credential_valid,
                                checked_at
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                            ON CONFLICT (server_id)
                            DO UPDATE SET
                                run_id = EXCLUDED.run_id,
                                trelix_installed = EXCLUDED.trelix_installed,
                                trelix_version = EXCLUDED.trelix_version,
                                status = EXCLUDED.status,
                                message = EXCLUDED.message,
                                credential_valid = EXCLUDED.credential_valid,
                                checked_at = NOW()
                            """,
                            (run_id, server_id, installed, version, status, message, credential_valid),
                        )

                    cur.execute(
                        "UPDATE deployment_runs SET completed_at = NOW() WHERE id = %s",
                        (run_id,),
                    )
    except Exception as exc:
        import traceback
        print(f"[DB] save run results failed: {exc}")
        traceback.print_exc()
    finally:
        if ENABLE_RESULT_EXPORTS:
            try:
                _export_secure_xlsx(platform, servers, per_ip_results)
            except Exception as exc:
                import traceback
                print(f"[XLSX] export failed: {exc}")
                traceback.print_exc()
        if conn:
            conn.close()


def _smtp_send(msg):
    """Send a MIMEMultipart message using LOGIN auth (required by smtp.mitel.com)."""
    import base64
    server = smtplib.SMTP(EMAIL_CONFIG["smtp_server"], EMAIL_CONFIG["smtp_port"], timeout=15)
    try:
        server.ehlo()
        server.starttls()
        server.ehlo()
        # Force LOGIN auth — server does not support PLAIN
        server.docmd("AUTH LOGIN")
        server.docmd(base64.b64encode(EMAIL_CONFIG["sender"].encode()).decode())
        code, resp = server.docmd(base64.b64encode(EMAIL_CONFIG["password"].encode()).decode())
        if code != 235:
            raise smtplib.SMTPAuthenticationError(code, resp)
        server.sendmail(EMAIL_CONFIG["sender"], EMAIL_CONFIG["recipient"], msg.as_string())
    finally:
        try:
            server.quit()
        except Exception:
            pass

# Track session results for email summary
session_results = {}
session_meta = {}


def send_summary_email(session_id, platform):
    """Send deployment failure report email - ONLY if there are failed VMs."""
    results = session_results.get(session_id, {})
    if not results:
        return

    config = get_platform_config(platform)
    platform_title = config["title"] if config else platform.upper()

    installed = [ip for ip, info in results.items() if info["status"] == "installed"]
    failed    = [ip for ip, info in results.items() if info["status"] in ("error", "warning")]
    total = len(results)

    # Only send email if there are failures
    if not failed:
        return

    now = datetime.now().strftime("%B %d, %Y  |  %H:%M")

    # Categorize failure reasons for clear display
    def get_failure_reason(msg):
        msg_lower = msg.lower()
        if "timed out" in msg_lower or "unreachable" in msg_lower:
            return "VM Turned Off / Unreachable", "#dc2626"
        elif "authentication" in msg_lower or "login failed" in msg_lower or "credential" in msg_lower:
            return "Invalid Credentials", "#9333ea"
        elif "refused" in msg_lower:
            return "Connection Refused (SSH down)", "#ea580c"
        elif "no route" in msg_lower or "offline" in msg_lower:
            return "Network Unreachable", "#0369a1"
        elif "root" in msg_lower:
            return "Root Login Failed", "#b45309"
        elif "upload" in msg_lower:
            return "Script Upload Failed", "#64748b"
        else:
            return "Error", "#dc2626"

    # Build failed server rows
    def fail_row(ip, info, idx):
        msg = info["message"]
        reason_label, reason_color = get_failure_reason(msg)
        row_bg = "#ffffff" if idx % 2 == 0 else "#fef2f2"
        return f"""<tr style="background-color:{row_bg};">
          <td style="padding:16px 20px;border-bottom:1px solid #fecaca;font-family:'Courier New',monospace;font-size:14px;font-weight:bold;color:#1f2937;">{ip}</td>
          <td style="padding:16px 20px;border-bottom:1px solid #fecaca;">
            <span style="display:inline-block;padding:4px 12px;border-radius:4px;font-size:11px;font-weight:bold;background-color:#fee2e2;border:2px solid #f87171;color:#991b1b;">FAILED</span>
          </td>
          <td style="padding:16px 20px;border-bottom:1px solid #fecaca;">
            <div style="font-size:13px;font-weight:bold;color:{reason_color};margin-bottom:4px;">{reason_label}</div>
            <div style="font-size:12px;color:#6b7280;">{msg}</div>
          </td>
        </tr>"""

    rows_html = "".join(
        fail_row(ip, results[ip], idx)
        for idx, ip in enumerate(sorted(failed, key=lambda x: [int(p) for p in x.split(".")]))
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<!--[if mso]><xml><o:OfficeDocumentSettings><o:AllowPNG/><o:PixelsPerInch>96</o:PixelsPerInch></o:OfficeDocumentSettings></xml><![endif]-->
<title>Trellix Failure Report</title>
<style type="text/css">
  @media only screen and (max-width: 620px) {{
    .email-container {{ width: 100% !important; }}
    .header-td {{ padding: 28px 20px 24px !important; }}
    .content-td {{ padding: 20px !important; }}
    .footer-td {{ padding: 16px 20px !important; }}
    .stat-table {{ width: 100% !important; }}
    .stat-td {{ display: block !important; width: 100% !important; padding: 6px 0 !important; }}
    .server-table {{ font-size: 12px !important; }}
    .server-table td {{ padding: 12px 10px !important; }}
    .title-text {{ font-size: 22px !important; }}
    .summary-box {{ padding: 10px 14px !important; }}
  }}
</style>
</head>
<body style="margin:0;padding:0;background-color:#f1f5f9;font-family:'Segoe UI',Helvetica,Arial,sans-serif;-webkit-text-size-adjust:100%;-ms-text-size-adjust:100%;">
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f1f5f9;padding:24px 10px;">
<tr><td align="center">
<table class="email-container" width="600" cellpadding="0" cellspacing="0" border="0" style="max-width:600px;width:100%;border-collapse:separate;border:1px solid #e2e8f0;">

  <!-- HEADER -->
  <tr><td class="header-td" style="background-color:#1e1b4b;padding:32px 36px 28px;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="font-size:18px;font-weight:bold;color:#ffffff;padding-bottom:6px;">
          &#128737; Trellix Manager
        </td>
        <td align="right">
          <span style="display:inline-block;background-color:#dc2626;border-radius:4px;padding:5px 12px;font-size:10px;font-weight:bold;color:#ffffff;letter-spacing:0.8px;">FAILURE ALERT</span>
        </td>
      </tr>
    </table>
    <div class="title-text" style="font-size:26px;font-weight:bold;color:#ffffff;margin-top:16px;">Deployment Failure Report</div>
    <div style="font-size:13px;color:#a5b4fc;margin-top:6px;font-weight:bold;">{platform_title} &mdash; {now}</div>
  </td></tr>

  <!-- RED ACCENT BAR -->
  <tr><td style="background-color:#dc2626;height:4px;font-size:0;line-height:0;">&nbsp;</td></tr>

  <!-- SUMMARY -->
  <tr><td class="content-td" style="background-color:#ffffff;padding:28px 36px 20px;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td class="stat-td" width="50%" style="padding-right:8px;vertical-align:top;">
          <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border:2px solid #dc2626;border-collapse:collapse;">
            <tr><td style="background-color:#dc2626;padding:10px 16px;text-align:center;">
              <span style="font-size:12px;font-weight:bold;color:#ffffff;letter-spacing:0.5px;">FAILED</span>
            </td></tr>
            <tr><td style="background-color:#fef2f2;padding:18px 16px;text-align:center;">
              <div style="font-size:36px;font-weight:bold;color:#dc2626;">{len(failed)}</div>
              <div style="font-size:11px;color:#991b1b;margin-top:4px;">out of {total} servers</div>
            </td></tr>
          </table>
        </td>
        <td class="stat-td" width="50%" style="padding-left:8px;vertical-align:top;">
          <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border:2px solid #16a34a;border-collapse:collapse;">
            <tr><td style="background-color:#16a34a;padding:10px 16px;text-align:center;">
              <span style="font-size:12px;font-weight:bold;color:#ffffff;letter-spacing:0.5px;">INSTALLED</span>
            </td></tr>
            <tr><td style="background-color:#f0fdf4;padding:18px 16px;text-align:center;">
              <div style="font-size:36px;font-weight:bold;color:#16a34a;">{len(installed)}</div>
              <div style="font-size:11px;color:#166534;margin-top:4px;">out of {total} servers</div>
            </td></tr>
          </table>
        </td>
      </tr>
    </table>
  </td></tr>

  <!-- FAILED SERVERS HEADING -->
  <tr><td class="content-td" style="background-color:#ffffff;padding:8px 36px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr><td style="border-top:2px solid #e5e7eb;padding-top:20px;">
        <span style="display:inline-block;background-color:#991b1b;border-radius:4px;padding:8px 16px;font-size:11px;font-weight:bold;letter-spacing:1px;color:#ffffff;">FAILED SERVERS &mdash; {len(failed)}</span>
      </td></tr>
    </table>
  </td></tr>

  <!-- FAILED SERVERS TABLE -->
  <tr><td class="content-td" style="background-color:#ffffff;padding:16px 36px 28px;">
    <table class="server-table" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;border:2px solid #fca5a5;">
      <thead>
        <tr style="background-color:#fef2f2;">
          <th style="padding:12px 20px;text-align:left;font-size:11px;font-weight:bold;color:#991b1b;border-bottom:2px solid #fca5a5;letter-spacing:0.8px;">SERVER IP</th>
          <th style="padding:12px 20px;text-align:left;font-size:11px;font-weight:bold;color:#991b1b;border-bottom:2px solid #fca5a5;letter-spacing:0.8px;">STATUS</th>
          <th style="padding:12px 20px;text-align:left;font-size:11px;font-weight:bold;color:#991b1b;border-bottom:2px solid #fca5a5;letter-spacing:0.8px;">FAILURE REASON</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </td></tr>

  <!-- FOOTER -->
  <tr><td class="footer-td" style="background-color:#1e1b4b;padding:18px 36px;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="font-size:12px;color:#a5b4fc;">
          <span style="font-weight:bold;color:#ffffff;">&#128737; Trellix Manager</span>
          <span style="color:#818cf8;">&nbsp;|&nbsp;</span>
          <span>Automated Security Deployment</span>
        </td>
        <td align="right" style="font-size:11px;color:#a5b4fc;">{now}</td>
      </tr>
    </table>
  </td></tr>

</table>
</td></tr>
</table>
</body>
</html>"""

    subject = f"[Trellix] {platform_title} — {len(failed)}/{total} FAILED"

    msg = MIMEMultipart("alternative")
    msg["From"] = EMAIL_CONFIG["sender"]
    msg["To"] = EMAIL_CONFIG["recipient"]
    msg["Subject"] = subject
    msg.attach(MIMEText(html, "html"))

    try:
        _smtp_send(msg)
        print(f"[Email] Summary sent to {EMAIL_CONFIG['recipient']}")
    except Exception as e:
        import traceback
        print(f"[Email] FAILED: {e}")
        print(traceback.format_exc())

    # Cleanup
    if session_id in session_results:
        del session_results[session_id]


def get_platform_config(platform):
    return PLATFORM_CONFIG.get(platform)


def get_value(entry, keys):
    for key in keys:
        value = entry.get(key, "")
        if value:
            return value
    return ""


def load_servers_from_file(filepath, platform='mxone'):
    """Read server list from uploaded Excel file."""
    config = get_platform_config(platform)
    if not config:
        raise ValueError(f"Unsupported platform: {platform}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    configured_root_password = str(config.get("root_password") or "").strip()
    force_configured_root = bool(configured_root_password) and platform != 'mxone'

    servers = []
    first_row_values = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Decide if row-1 is a header row; otherwise treat sheet as data-only with fixed columns.
    expected_headers = set(
        config["headers"]["ip"]
        + config["headers"]["username"]
        + config["headers"]["password"]
        + config["headers"].get("root_password", [])
    )
    has_header = any(h in expected_headers for h in first_row_values)

    if has_header:
        if platform == 'mxone':
            required_sets = [
                set(config["headers"]["ip"]),
                set(config["headers"]["username"]),
                set(config["headers"]["password"]),
                set(config["headers"].get("root_password", [])),
            ]
            missing_any = any(not (set(first_row_values) & req) for req in required_sets)
            if missing_any:
                raise ValueError("MxOne file must include headers: IP, Putty Username, Putty Password, Root Password")

        headers = first_row_values
        data_rows = ws.iter_rows(min_row=2, values_only=True)
        for row in data_rows:
            if not row or not row[0]:
                continue
            entry = dict(zip(headers, [str(v).strip() if v else "" for v in row]))

            server = {
                "ip": get_value(entry, config["headers"]["ip"]),
                "admin_username": get_value(entry, config["headers"]["username"]),
                "admin_password": get_value(entry, config["headers"]["password"]),
                "root_password": configured_root_password if force_configured_root else config.get("root_password", get_value(entry, config["headers"].get("root_password", []))),
            }

            has_basic = server["ip"] and server["admin_username"] and server["admin_password"]
            has_root_for_mxone = (platform != 'mxone') or bool(server.get("root_password"))
            if has_basic and has_root_for_mxone:
                servers.append(server)
    else:
        # Headerless format: A=IP, B=username, C=password, D(optional)=root password.
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue

            ip = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            username = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            password = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            root_pw = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""

            has_basic = ip and username and password
            has_root_for_mxone = (platform != 'mxone') or bool(root_pw)
            if has_basic and has_root_for_mxone:
                servers.append({
                    "ip": ip,
                    "admin_username": username,
                    "admin_password": password,
                    "root_password": configured_root_password if force_configured_root else (root_pw or configured_root_password),
                })

    if platform == 'mxone' and not servers:
        raise ValueError("MxOne file must contain valid rows with: IP, Putty Username, Putty Password, Root Password")

    return servers


def send_event(session_id, ip, status, message, details=None):
    """Send an event to the client."""
    meta = session_meta.get(session_id, {})
    platform_name = str(meta.get("platform") or "")
    _append_audit_event(
        session_id=session_id,
        platform=platform_name,
        ip=ip,
        phase="event",
        result=status,
        message=message,
        details=details or "",
    )
    log_structured(
        "event",
        session_id=session_id,
        vm_ip=ip,
        platform=platform_name,
        phase="event",
        result=status,
        duration_ms=0,
        message=str(message or ""),
    )
    event = {
        "ip": ip,
        "status": status,
        "message": message,
        "details": details or ""
    }
    if session_id in event_queues:
        event_queues[session_id].put(event)


def strip_ansi_codes(text):
    """Remove ANSI escape codes and control characters from text."""
    import re
    # Remove ANSI escape sequences (colors, cursor moves, etc.)
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    text = ansi_escape.sub('', text)
    # Remove OSC hyperlink sequences: ]8;;...\\...  ]8;;\\
    text = re.sub(r'\]8;;[^\x1b]*(?:\x1b\\|\\)', '', text)
    # Remove common terminal control characters
    text = text.replace('\x1b[?1h=', '')
    text = text.replace('\x1b[K', '')
    text = text.replace('\x1b(B', '')
    text = re.sub(r'\[\?1h=?', '', text)
    text = re.sub(r'\[END\].*', '', text)
    # Remove residual 8;;...8;; hyperlink markers
    text = re.sub(r'8;;[^\s]*8;;', '', text)
    # Remove trailing > or ~ from pager line wrapping
    text = re.sub(r'[>~]$', '', text)
    # Remove [7m...[m escape leftovers
    text = re.sub(r'\[\d*m', '', text)
    return text.strip()


def _get_linux_xagt_version(ssh):
    """Best-effort Linux version lookup via rpm."""
    try:
        stdin, stdout, stderr = ssh.exec_command("rpm -q xagt 2>&1 || true", timeout=8)
        out = (stdout.read() or b"").decode(errors='replace').strip()
        if out and "not installed" not in out.lower() and "command not found" not in out.lower():
            return out.splitlines()[-1].strip()
    except Exception:
        pass
    return ""


def _get_windows_trellix_version(session):
    """Best-effort Windows version lookup from xagt executable metadata."""
    cmd = r"""
$paths = @(
  'C:\Program Files\McAfee\Agent\x64\xagt.exe',
  'C:\Program Files\McAfee\Agent\xagt.exe'
)
foreach ($p in $paths) {
  if (Test-Path $p) {
    $v = (Get-Item $p).VersionInfo.ProductVersion
    if ($v) { Write-Output ('VERSION:' + $v); exit 0 }
  }
}
Write-Output 'VERSION:'
"""
    try:
        res = session.run_ps(cmd)
        out = ((res.std_out or b"") + (res.std_err or b"")).decode(errors="replace")
        for line in out.splitlines():
            if line.startswith("VERSION:"):
                return line.split("VERSION:", 1)[1].strip()
    except Exception:
        pass
    return ""


def _installed_message_with_version(version):
    """Build installed status message with optional version suffix."""
    ver = str(version or "").strip()
    if ver:
        return f"Trellix is INSTALLED and Running ({ver})"
    return "Trellix is INSTALLED and Running"


_WINDOWS_SCP_CACHE = {}
_WINDOWS_SCP_CACHE_LOCK = threading.Lock()


def _is_windows_scp_source_configured(config):
    """Return True when Linux source settings needed for SCP download are present."""
    return all([
        str(config.get("scp_source_host") or "").strip(),
        str(config.get("scp_source_user") or "").strip(),
        str(config.get("scp_source_pass") or "").strip(),
        str(config.get("scp_source_path") or "").strip(),
    ])


def _windows_scp_local_zip(config):
    """Download source ZIP from Linux host once and return local temp path."""
    host = str(config.get("scp_source_host") or "").strip()
    user = str(config.get("scp_source_user") or "").strip()
    password = str(config.get("scp_source_pass") or "").strip()
    remote_path = str(config.get("scp_source_path") or "").strip()
    port = int(config.get("scp_source_port") or 22)

    if not host or not user or not password or not remote_path:
        raise RuntimeError("SCP source is not configured")

    cache_key = f"{host}:{port}|{user}|{remote_path}"
    with _WINDOWS_SCP_CACHE_LOCK:
        cached = _WINDOWS_SCP_CACHE.get(cache_key)
        if cached and os.path.exists(cached):
            return cached

    source_name = os.path.basename(remote_path.rstrip("/\\")) or "windows_agent.zip"
    local_path = os.path.join(tempfile.gettempdir(), f"trelix_{abs(hash(cache_key))}_{source_name}")

    if not os.path.exists(local_path):
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            ssh.connect(
                hostname=host,
                port=port,
                username=user,
                password=password,
                timeout=30,
                allow_agent=False,
                look_for_keys=False,
            )
            sftp = ssh.open_sftp()
            try:
                try:
                    sftp.stat(remote_path)
                except FileNotFoundError:
                    raise RuntimeError(f"Linux source ZIP not found: {remote_path}")
                except OSError as exc:
                    raise RuntimeError(f"Cannot access Linux source ZIP: {remote_path} ({exc})")
                sftp.get(remote_path, local_path)
            finally:
                sftp.close()
        finally:
            ssh.close()

    with _WINDOWS_SCP_CACHE_LOCK:
        _WINDOWS_SCP_CACHE[cache_key] = local_path
    return local_path


def _sftp_mkdir_p(sftp, path):
    """Create remote directory path recursively if it does not exist."""
    normalized = str(path or "").replace('\\', '/').rstrip('/')
    if not normalized:
        return
    parts = [p for p in normalized.split('/') if p]
    if not parts:
        return

    current = ''
    start_index = 0

    # Windows SFTP paths can begin with a drive root like C:. Never mkdir drive roots.
    if len(parts[0]) == 2 and parts[0][1] == ':':
        current = f"/{parts[0]}"
        start_index = 1
        try:
            sftp.stat(current)
        except Exception as exc:
            raise RuntimeError(f"SFTP drive root not accessible: {current} ({exc})")

    for part in parts[start_index:]:
        current = f"{current}/{part}" if current else f"/{part}"
        try:
            sftp.stat(current)
        except Exception:
            try:
                sftp.mkdir(current)
            except Exception as exc:
                raise RuntimeError(f"SFTP cannot create directory: {current} ({exc})")


def _scp_upload_windows_zip(ip, username, password, local_zip_path, target_dir, package_basename=""):
    """Upload local ZIP to Windows target directory over SSH/SFTP."""
    requested = str(package_basename or "").strip()
    if requested:
        name = requested if requested.lower().endswith('.zip') else f"{requested}.zip"
    else:
        name = os.path.basename(local_zip_path)

    configured_native = str(target_dir or r"C:\Users\Administrator\Downloads").replace('\\', '/').rstrip('/')
    remote_dir_native_candidates = []

    def _add_remote_dir(p):
        val = str(p or "").replace('\\', '/').rstrip('/')
        if val and val not in remote_dir_native_candidates:
            remote_dir_native_candidates.append(val)

    _add_remote_dir(configured_native)
    _add_remote_dir(f"C:/Users/{username}/Downloads")
    _add_remote_dir(r"C:/Users/Administrator/Downloads")
    _add_remote_dir(r"C:/Windows/Temp")

    remote_dir_candidates = []
    for native in remote_dir_native_candidates:
        if not native.startswith('/'):
            remote_dir_candidates.append('/' + native)
        remote_dir_candidates.append(native)

    put_errors = []
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(
            hostname=ip,
            port=22,
            username=username,
            password=password,
            timeout=30,
            allow_agent=False,
            look_for_keys=False,
        )
        sftp = ssh.open_sftp()
        try:
            success_remote_path = ""
            for remote_dir in remote_dir_candidates:
                try:
                    _sftp_mkdir_p(sftp, remote_dir)
                    remote_sftp_path = f"{remote_dir}/{name}"
                    sftp.put(local_zip_path, remote_sftp_path)
                    sftp.stat(remote_sftp_path)
                    success_remote_path = remote_sftp_path
                    break
                except Exception as exc:
                    put_errors.append(f"{remote_dir}: {exc}")
            else:
                tried = " | ".join(put_errors) or "unknown SFTP error"
                raise RuntimeError(f"SFTP upload failed for target directory '{target_dir}'. Tried: {tried}")

            if not success_remote_path:
                raise RuntimeError("SFTP upload reported success without a resolved remote path")

            # Normalize path returned to WinRM PowerShell-friendly Windows path.
            resolved_native = success_remote_path.lstrip('/').replace('/', '\\')
        finally:
            sftp.close()
    except Exception as exc:
        raise RuntimeError(f"Windows SSH/SFTP upload failed for {ip} with user '{username}': {exc}")
    finally:
        ssh.close()

    return resolved_native


def _winrm_upload_windows_zip(session, local_zip_path, target_dir, package_basename=""):
    """Upload local ZIP to Windows over WinRM as a fallback when SSH/SFTP is unavailable."""
    if not os.path.exists(local_zip_path):
        raise RuntimeError(f"Local ZIP not found for WinRM upload: {local_zip_path}")

    requested = str(package_basename or "").strip()
    if requested:
        name = requested if requested.lower().endswith('.zip') else f"{requested}.zip"
    else:
        name = os.path.basename(local_zip_path)

    target_dir = str(target_dir or r"C:\Users\Administrator\Downloads").replace('/', '\\').rstrip('\\')
    remote_path = f"{target_dir}\\{name}"
    target_dir_escaped = target_dir.replace("'", "''")
    remote_path_escaped = remote_path.replace("'", "''")

    init_cmd = f"""
$dir = '{target_dir_escaped}'
$path = '{remote_path_escaped}'
New-Item -Path $dir -ItemType Directory -Force | Out-Null
[System.IO.File]::WriteAllBytes($path, [byte[]]@())
Write-Output ('WINRM_UPLOAD_PATH:' + $path)
"""
    init_res = session.run_ps(init_cmd)
    init_out = ((init_res.std_out or b"") + (init_res.std_err or b"")).decode(errors="replace")
    if "WINRM_UPLOAD_PATH:" not in init_out:
        raise RuntimeError(f"Failed to initialize WinRM upload path: {init_out.strip() or 'unknown error'}")

    # Keep each inline base64 payload small to avoid WinRM/PowerShell command length faults.
    chunk_size = int(os.getenv("WINDOWS_WINRM_UPLOAD_CHUNK_BYTES", "4096") or "4096")
    if chunk_size < 1024:
        chunk_size = 1024
    if chunk_size > 8192:
        chunk_size = 8192
    with open(local_zip_path, "rb") as fh:
        while True:
            chunk = fh.read(chunk_size)
            if not chunk:
                break
            chunk_b64 = base64.b64encode(chunk).decode("ascii")
            chunk_cmd = f"""
$path = '{remote_path_escaped}'
$bytes = [Convert]::FromBase64String('{chunk_b64}')
$fs = [System.IO.File]::Open($path, [System.IO.FileMode]::Append, [System.IO.FileAccess]::Write, [System.IO.FileShare]::Read)
$fs.Write($bytes, 0, $bytes.Length)
$fs.Close()
Write-Output 'WINRM_CHUNK_OK'
"""
            chunk_res = session.run_ps(chunk_cmd)
            chunk_out = ((chunk_res.std_out or b"") + (chunk_res.std_err or b"")).decode(errors="replace")
            if "WINRM_CHUNK_OK" not in chunk_out:
                raise RuntimeError(f"WinRM chunk upload failed: {chunk_out.strip() or 'unknown error'}")

    verify_cmd = f"""
$path = '{remote_path_escaped}'
if (-not (Test-Path $path)) {{ Write-Output 'WINRM_UPLOAD_MISSING'; exit 1 }}
$len = (Get-Item -Path $path).Length
Write-Output ('WINRM_UPLOAD_SIZE:' + $len)
"""
    verify_res = session.run_ps(verify_cmd)
    verify_out = ((verify_res.std_out or b"") + (verify_res.std_err or b"")).decode(errors="replace")
    if "WINRM_UPLOAD_SIZE:" not in verify_out:
        raise RuntimeError(f"WinRM upload verification failed: {verify_out.strip() or 'unknown error'}")

    remote_size = 0
    for line in verify_out.splitlines():
        if line.startswith("WINRM_UPLOAD_SIZE:"):
            try:
                remote_size = int(line.split("WINRM_UPLOAD_SIZE:", 1)[1].strip())
            except Exception:
                remote_size = 0
            break

    local_size = os.path.getsize(local_zip_path)
    if remote_size != local_size:
        raise RuntimeError(f"WinRM upload size mismatch: remote={remote_size}, local={local_size}")

    return remote_path


def _parse_unc_root(unc_path):
    """Return UNC root like \\server\share from a full UNC path."""
    raw = str(unc_path or "").strip()
    if not raw.startswith('\\\\'):
        return ""
    parts = raw.lstrip('\\').split('\\')
    if len(parts) < 2:
        return ""
    return f"\\\\{parts[0]}\\{parts[1]}"


def _windows_unc_candidates(config):
    """Build resilient UNC candidates for Windows installer source."""
    raw_unc = str((config or {}).get("installer_unc") or "").strip()
    candidates = []

    def _add(path):
        p = str(path or "").strip()
        if p and p not in candidates:
            candidates.append(p)

    if raw_unc:
        _add(raw_unc)
        if "\\trelix_credentails\\" in raw_unc.lower():
            _add(raw_unc.replace("\\trelix_credentails\\", "\\trelix_credentials\\"))
        if "\\trelix_credentials\\" in raw_unc.lower():
            _add(raw_unc.replace("\\trelix_credentials\\", "\\trelix_credentails\\"))

    root = _parse_unc_root(raw_unc) if raw_unc else f"\\\\{SMB_SERVER}\\{SMB_SHARE}"
    for sub in _candidate_share_dirs("windows"):
        sub_win = sub.replace('/', '\\')
        _add(f"{root}\\{sub_win}")

    return candidates


def _test_unc_path_for_installer(session, unc_path):
    """Test if a UNC path contains installer files (MSI/EXE or ZIP with installer).
    
    Returns (has_installer, file_count, details_message)
    """
    test_cmd = f"""
$path = '{str(unc_path or "").replace("'", "''")}'
if (-not (Test-Path $path)) {{ Write-Output 'PATH_NOT_FOUND'; exit 1 }}
$items = @(Get-ChildItem -Path $path -File -ErrorAction SilentlyContinue)
$installers = @($items | Where-Object {{ $_.Extension -ieq '.msi' -or $_.Extension -ieq '.exe' }})
$zips = @($items | Where-Object {{ $_.Extension -ieq '.zip' }})
Write-Output "ITEMS_COUNT:$($items.Count)"
Write-Output "INSTALLERS_COUNT:$($installers.Count)"
Write-Output "ZIPS_COUNT:$($zips.Count)"
if ($installers.Count -gt 0) {{ Write-Output 'HAS_INSTALLER' }}
if ($zips.Count -gt 0) {{ Write-Output 'HAS_ZIP' }}
"""
    try:
        res = session.run_ps(test_cmd)
        out = ((res.std_out or b"") + (res.std_err or b"")).decode(errors="replace")
        
        has_installer = 'HAS_INSTALLER' in out or 'HAS_ZIP' in out
        items_count = 0
        installers_count = 0
        
        for line in out.splitlines():
            if line.startswith('ITEMS_COUNT:'):
                items_count = int(line.split(':')[1])
            elif line.startswith('INSTALLERS_COUNT:'):
                installers_count = int(line.split(':')[1])
        
        msg = f"({items_count} total files, {installers_count} installers)"
        return has_installer, items_count, msg
    except Exception as exc:
        return False, 0, str(exc)


def _ensure_windows_sshd_for_scp(session):
        """Ensure OpenSSH server and firewall rule exist for SCP transfers."""
        ps = r"""
$svc = Get-Service -Name sshd -ErrorAction SilentlyContinue
if ($svc) {
    if ($svc.Status -ne 'Running') {
        Start-Service sshd -ErrorAction SilentlyContinue
        Set-Service -Name sshd -StartupType Automatic -ErrorAction SilentlyContinue
        $svc = Get-Service -Name sshd -ErrorAction SilentlyContinue
    }
    if ($svc -and $svc.Status -eq 'Running') { Write-Output 'SSHD_READY'; exit 0 }
}

$cap = Get-WindowsCapability -Online -Name OpenSSH.Server~~~~0.0.1.0 -ErrorAction SilentlyContinue
if ($cap -and $cap.State -ne 'Installed') {
    Add-WindowsCapability -Online -Name OpenSSH.Server~~~~0.0.1.0 | Out-Null
}

Start-Service sshd -ErrorAction SilentlyContinue
Set-Service -Name sshd -StartupType Automatic -ErrorAction SilentlyContinue

$rule = Get-NetFirewallRule -Name sshd -ErrorAction SilentlyContinue
if (-not $rule) {
    New-NetFirewallRule -Name sshd -DisplayName 'OpenSSH Server (sshd)' -Enabled True -Direction Inbound -Protocol TCP -Action Allow -LocalPort 22 | Out-Null
}

$svc = Get-Service -Name sshd -ErrorAction SilentlyContinue
if ($svc -and $svc.Status -eq 'Running') {
    Write-Output 'SSHD_READY'
} else {
    Write-Output 'SSHD_NOT_READY'
}
"""

        res = session.run_ps(ps)
        out = ((res.std_out or b"") + (res.std_err or b"")).decode(errors="replace")
        return ("SSHD_READY" in out, out.strip())


def _verify_linux_platform_identity(ssh, platform):
    """Verify the target Linux host matches the selected deployment platform."""
    checks = {
        "mivb": {
            "cmd": "if [ -d /var/mivb ]; then echo 'MiVB Detected'; else echo 'Not MiVB'; fi",
            "ok": lambda out: "mivb detected" in out.lower(),
            "expected": "MiVB Manifest",
        },
        "mxone": {
            "cmd": "(ts_about 2>&1 || /usr/eri/bin/ts_about 2>&1 || true)",
            "ok": lambda out: ("mivoice mx-one" in out.lower()) or ("mx-one" in out.lower()),
            "expected": "MX-ONE ts_about signature",
        },
        "miv5000": {
            "cmd": "if [ -d /opt/a5000 ]; then echo 'MiV5000 Detected'; else echo 'Not MiV5000'; fi",
            "ok": lambda out: "miv5000 detected" in out.lower(),
            "expected": "/opt/a5000 present",
        },
        "mbg": {
            "cmd": "if systemctl list-unit-files | grep -qi mbg; then echo 'MBG Detected'; else echo 'Not MBG'; fi",
            "ok": lambda out: "mbg detected" in out.lower(),
            "expected": "MBG unit files present",
        },
    }

    cfg = checks.get(platform)
    if not cfg:
        return True, "No platform identity check required"

    try:
        stdin, stdout, stderr = ssh.exec_command(cfg["cmd"], timeout=10)
        out = (stdout.read() or b"").decode(errors="replace")
        err = (stderr.read() or b"").decode(errors="replace")
        combined = f"{out}\n{err}".strip()
        if cfg["ok"](combined):
            return True, "Platform identity verified"
        return False, (
            f"Platform mismatch: selected '{platform}' but expected {cfg['expected']} was not detected. "
            f"Command output: {combined[:250] or 'empty'}"
        )
    except Exception as exc:
        return False, f"Platform verification failed: {exc}"


def process_windows_server(session_id, server, config):
    """Process a Windows server via WinRM: check Trellix service and optionally install."""
    ip = server["ip"]
    admin_username = server["admin_username"]
    admin_password = server["admin_password"]

    send_event(session_id, ip, "connecting", "Connecting...")

    endpoint = f"http://{ip}:{config.get('winrm_port', 5985)}/wsman"
    transport = config.get("winrm_transport", "ntlm")

    try:
        session = winrm.Session(endpoint, auth=(admin_username, admin_password), transport=transport)
    except Exception as e:
        send_event(session_id, ip, "error", "Connection Failed", str(e))
        session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": f"WinRM session failed: {e}"}
        return

    send_event(session_id, ip, "progress", "WinRM Login Successful")

    # Pre-check services expected in your environment.
    precheck_cmd = r"""
$winrm = Get-Service -Name WinRM -ErrorAction SilentlyContinue
$sshd = Get-Service -Name sshd -ErrorAction SilentlyContinue
if ($winrm -and $winrm.Status -eq 'Running') { Write-Output 'WINRM_OK' } else { Write-Output 'WINRM_NOT_RUNNING' }
if ($sshd -and $sshd.Status -eq 'Running') { Write-Output 'SSHD_OK' } else { Write-Output 'SSHD_NOT_RUNNING' }
"""

    try:
        precheck_res = session.run_ps(precheck_cmd)
        precheck_out = ((precheck_res.std_out or b"") + (precheck_res.std_err or b"")).decode(errors="replace")
        if "WINRM_OK" in precheck_out:
            send_event(session_id, ip, "log", "WinRM service is running.")
        else:
            send_event(session_id, ip, "warning", "WinRM service is not running", "Run: winrm quickconfig and open firewall port 5985.")

        if "SSHD_OK" in precheck_out:
            send_event(session_id, ip, "log", "OpenSSH service is running.")
        else:
            send_event(session_id, ip, "warning", "OpenSSH service is not running", "Install OpenSSH Server and start sshd.")
    except Exception:
        # Do not fail deployment for pre-check read errors.
        pass

    send_event(session_id, ip, "progress", "Checking Trellix status...")

    status_cmd = r"""
$svc = Get-Service -Name xagt -ErrorAction SilentlyContinue
if (-not $svc) {
    $svc = Get-Service -Name masvc -ErrorAction SilentlyContinue
}
if ($svc -and $svc.Status -eq 'Running') {
    Write-Output ('TRELLIX_RUNNING:' + $svc.Name)
} elseif ($svc) {
    Write-Output ('TRELLIX_INSTALLED_STOPPED:' + $svc.Name)
} else {
    Write-Output 'TRELLIX_NOT_INSTALLED'
}
"""

    try:
        status_res = session.run_ps(status_cmd)
    except Exception as e:
        send_event(session_id, ip, "error", "Connection Failed", str(e))
        session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": f"WinRM command failed: {e}"}
        return

    output = (status_res.std_out or b"").decode(errors="replace")
    err = (status_res.std_err or b"").decode(errors="replace")
    combined = f"{output}\n{err}".strip()

    detected_service = "xagt" if "xagt" in combined.lower() else ("masvc" if "masvc" in combined.lower() else "xagt")

    if "TRELLIX_RUNNING" in combined:
        send_event(session_id, ip, "log", f"Detected Trellix service: {detected_service}")
        win_version = _get_windows_trellix_version(session)
        send_event(session_id, ip, "installed", _installed_message_with_version(win_version))
        session_results.setdefault(session_id, {})[ip] = {
            "status": "installed",
            "message": "Already installed and running (Windows)",
            "version": win_version,
        }
        return

    if "TRELLIX_INSTALLED_STOPPED" in combined:
        send_event(session_id, ip, "log", f"Trellix service ({detected_service}) exists but is not running. Trying to start service...")
        start_res = session.run_ps(f"Start-Service -Name {detected_service} -ErrorAction Stop; Write-Output 'SERVICE_STARTED'")
        start_out = ((start_res.std_out or b"") + (start_res.std_err or b"")).decode(errors="replace")
        if "SERVICE_STARTED" in start_out:
            win_version = _get_windows_trellix_version(session)
            send_event(session_id, ip, "installed", _installed_message_with_version(win_version))
            session_results.setdefault(session_id, {})[ip] = {
                "status": "installed",
                "message": "Service started successfully",
                "version": win_version,
            }
            return
        send_event(session_id, ip, "warning", "Trellix installed but service could not be started", start_out.strip() or "Unknown service start issue")
        session_results.setdefault(session_id, {})[ip] = {"status": "warning", "message": "Service start failed"}
        return

    installer_unc = ""
    installer_unc_candidates = []
    if bool(config.get("scp_enabled")):
        if _is_windows_scp_source_configured(config):
            try:
                send_event(
                    session_id,
                    ip,
                    "log",
                    f"Using Linux SCP source: {config.get('scp_source_user')}@{config.get('scp_source_host')}:{config.get('scp_source_path')}",
                )
                if bool(config.get("scp_auto_setup_sshd", True)):
                    send_event(session_id, ip, "progress", "Checking OpenSSH (sshd) status for SCP...")
                    ssh_ready, ssh_msg = _ensure_windows_sshd_for_scp(session)
                    if ssh_ready:
                        send_event(session_id, ip, "log", "OpenSSH service is installed and running.")
                    else:
                        send_event(session_id, ip, "warning", "OpenSSH setup could not be fully verified", ssh_msg)
                send_event(session_id, ip, "progress", r"Transferring package to C:\Users\Administrator\Downloads (SCP with WinRM fallback)...")
                session.run_ps(r"New-Item -Path 'C:\Users\Administrator\Downloads' -ItemType Directory -Force | Out-Null")
                local_zip = _windows_scp_local_zip(config)
                installer_unc = _scp_upload_windows_zip(
                    ip=ip,
                    username=admin_username,
                    password=admin_password,
                    local_zip_path=local_zip,
                    target_dir=config.get("scp_target_dir") or r"C:\Users\Administrator\Downloads",
                    package_basename=config.get("package_basename") or "IMAGE_HX_AGENT_WIN_36.30.37",
                )
                send_event(session_id, ip, "log", f"SCP transfer completed: {installer_unc}")
                installer_unc_candidates = [installer_unc]
            except Exception as exc:
                # SSH/SFTP can be blocked on some hosts even when WinRM is reachable.
                # Fallback to WinRM file upload path before optional UNC fallback.
                try:
                    send_event(session_id, ip, "warning", "SCP transfer failed, trying WinRM upload fallback", str(exc))
                    local_zip = local_zip if 'local_zip' in locals() else _windows_scp_local_zip(config)
                    installer_unc = _winrm_upload_windows_zip(
                        session=session,
                        local_zip_path=local_zip,
                        target_dir=config.get("scp_target_dir") or r"C:\Users\Administrator\Downloads",
                        package_basename=config.get("package_basename") or "IMAGE_HX_AGENT_WIN_36.30.37",
                    )
                    send_event(session_id, ip, "log", f"WinRM file upload completed: {installer_unc}")
                    installer_unc_candidates = [installer_unc]
                except Exception as winrm_upload_exc:
                    if bool(config.get("scp_unc_fallback", False)):
                        send_event(session_id, ip, "warning", "WinRM upload fallback failed, trying UNC fallback", str(winrm_upload_exc))
                    else:
                        send_event(session_id, ip, "error", "SCP transfer failed", f"{exc} | WinRM fallback failed: {winrm_upload_exc}")
                        session_results.setdefault(session_id, {})[ip] = {
                            "status": "error",
                            "message": f"SCP transfer failed: {exc}; WinRM fallback failed: {winrm_upload_exc}",
                        }
                        return
        else:
            send_event(session_id, ip, "error", "SCP source not configured", "Set WINDOWS_SCP_SOURCE_HOST, WINDOWS_SCP_SOURCE_USER, WINDOWS_SCP_SOURCE_PASS, and WINDOWS_SCP_SOURCE_PATH")
            session_results.setdefault(session_id, {})[ip] = {
                "status": "error",
                "message": "SCP source is not fully configured",
            }
            return

    if not installer_unc:
        installer_unc_candidates = _windows_unc_candidates(config)
        send_event(session_id, ip, "progress", f"Validating {len(installer_unc_candidates)} installer source candidate(s)...")
        
        # Test each candidate to find one with installer files
        for cand in installer_unc_candidates:
            has_inst, count, msg = _test_unc_path_for_installer(session, cand)
            if has_inst:
                installer_unc = cand
                send_event(session_id, ip, "log", f"Found valid installer source at {cand} {msg}")
                break
            else:
                send_event(session_id, ip, "log", f"Checked {cand} - no installers found {msg}")
        
        if not installer_unc and installer_unc_candidates:
            # None of the candidates had installers - provide detailed error
            attempted = " | ".join(installer_unc_candidates)
            send_event(session_id, ip, "error", "Required File Missing", 
                f"No installer files found in any candidate paths: {attempted}. "
                f"Ensure installer files (MSI/EXE or ZIP containing them) are present at one of these locations.")
            session_results.setdefault(session_id, {})[ip] = {
                "status": "error",
                "message": f"Installer files not found at: {attempted}"
            }
            return

    send_event(session_id, ip, "progress", "Trellix not installed. Starting installation...")
    installer_unc_escaped = installer_unc.replace("'", "''")

    send_event(session_id, ip, "progress", "Preparing installer files in Downloads...")
    stage_cmd = f"""
$source = '{installer_unc_escaped}'
if (-not (Test-Path $source)) {{ Write-Output ('INSTALLER_SOURCE_NOT_FOUND:' + $source); exit 1 }}
$downloadsRoot = 'C:\\Users\\Administrator\\Downloads'
New-Item -Path $downloadsRoot -ItemType Directory -Force | Out-Null
$localWorkDir = Join-Path $downloadsRoot ('trellix_pkg_' + [guid]::NewGuid().ToString('N'))
New-Item -Path $localWorkDir -ItemType Directory -Force | Out-Null
$item = Get-Item -Path $source -ErrorAction Stop
if ($item.PSIsContainer) {{
  Copy-Item -Path (Join-Path $source '*') -Destination $localWorkDir -Recurse -Force -ErrorAction Stop
}} else {{
  Copy-Item -Path $source -Destination (Join-Path $localWorkDir $item.Name) -Force -ErrorAction Stop
}}
$scanDir = $localWorkDir
$zip = Get-ChildItem -Path $localWorkDir -Recurse -File -Filter *.zip -ErrorAction SilentlyContinue | Select-Object -First 1
if ($zip) {{
    $extractDir = Join-Path $localWorkDir ([IO.Path]::GetFileNameWithoutExtension($zip.Name))
  New-Item -Path $extractDir -ItemType Directory -Force | Out-Null
  Expand-Archive -Path $zip.FullName -DestinationPath $extractDir -Force
  $scanDir = $extractDir
}}
$installer = Get-ChildItem -Path $scanDir -Recurse -File -ErrorAction SilentlyContinue |
  Where-Object {{ $_.Extension -ieq '.msi' -or $_.Extension -ieq '.exe' }} |
  Select-Object -First 1
if (-not $installer) {{ Write-Output 'INSTALLER_NOT_FOUND'; exit 1 }}
$itype = if ($installer.Extension -ieq '.msi') {{ 'MSI' }} else {{ 'EXE' }}
Write-Output ('INSTALLER_SOURCE_USED:' + $source)
Write-Output ('LOCAL_WORK_DIR:' + $localWorkDir)
Write-Output ('INSTALLER_PATH:' + $installer.FullName)
Write-Output ('INSTALLER_TYPE:' + $itype)
"""

    stage_res = session.run_ps(stage_cmd)
    install_out = ((stage_res.std_out or b"") + (stage_res.std_err or b"")).decode(errors="replace")

    if "INSTALLER_SOURCE_NOT_FOUND" in install_out:
        send_event(session_id, ip, "error", "Required File Missing", f"Installer source path not accessible: {installer_unc}")
        session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": "Installer source not accessible"}
        return

    if "INSTALLER_NOT_FOUND" in install_out:
        send_event(session_id, ip, "error", "Required File Missing", f"No MSI/EXE installer found in {installer_unc}")
        session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": "Installer not found in source"}
        return

    if "LOCAL_WORK_DIR:" in install_out:
        try:
            source_used_line = [ln for ln in install_out.splitlines() if ln.startswith("INSTALLER_SOURCE_USED:")]
            if source_used_line:
                source_used = source_used_line[-1].split("INSTALLER_SOURCE_USED:", 1)[1].strip()
                if source_used:
                    send_event(session_id, ip, "log", f"Installer source resolved: {source_used}")
            local_dir_line = [ln for ln in install_out.splitlines() if ln.startswith("LOCAL_WORK_DIR:")][-1]
            local_dir = local_dir_line.split("LOCAL_WORK_DIR:", 1)[1].strip()
            if local_dir:
                send_event(session_id, ip, "log", f"Installer transferred to local path: {local_dir}")
        except Exception:
            pass

    installer_path = ""
    installer_type = ""
    try:
        for ln in install_out.splitlines():
            if ln.startswith("INSTALLER_PATH:"):
                installer_path = ln.split("INSTALLER_PATH:", 1)[1].strip()
            elif ln.startswith("INSTALLER_TYPE:"):
                installer_type = ln.split("INSTALLER_TYPE:", 1)[1].strip().upper()
    except Exception:
        installer_path = ""
        installer_type = ""

    if not installer_path or installer_type not in ("MSI", "EXE"):
        send_event(session_id, ip, "error", "Installation Failed", install_out.strip() or "Failed to resolve installer path")
        session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": "Installer resolution failed"}
        return

    installer_path_escaped = installer_path.replace("'", "''")
    send_event(session_id, ip, "progress", f"Installing Trellix package ({installer_type})...")
    install_run_cmd = f"""
$installer = '{installer_path_escaped}'
$itype = '{installer_type}'
$exitCode = 1
if ($itype -eq 'MSI') {{
  $proc = Start-Process msiexec.exe -ArgumentList "/i `"$installer`" /qn /norestart" -Wait -PassThru
  $exitCode = $proc.ExitCode
}} else {{
  $proc = Start-Process -FilePath $installer -ArgumentList '/quiet /norestart' -Wait -PassThru
  $exitCode = $proc.ExitCode
}}
if ($exitCode -eq 0 -or $exitCode -eq 3010) {{
  $svc = Get-Service -Name xagt -ErrorAction SilentlyContinue
  if (-not $svc) {{ $svc = Get-Service -Name masvc -ErrorAction SilentlyContinue }}
  if ($svc) {{ Start-Service -Name $svc.Name -ErrorAction SilentlyContinue }}
  $svc = Get-Service -Name xagt -ErrorAction SilentlyContinue
  if (-not $svc) {{ $svc = Get-Service -Name masvc -ErrorAction SilentlyContinue }}
  if ($svc -and $svc.Status -eq 'Running') {{ Write-Output 'INSTALL_OK_RUNNING' }} else {{ Write-Output 'INSTALL_OK_NOT_RUNNING' }}
}} else {{
  Write-Output ('INSTALL_EXIT:' + $exitCode)
  Write-Output 'INSTALL_FAILED'
}}
"""
    run_res = session.run_ps(install_run_cmd)
    install_out = ((run_res.std_out or b"") + (run_res.std_err or b"")).decode(errors="replace")

    send_event(session_id, ip, "progress", "Finalizing service status...")

    if "INSTALL_OK_RUNNING" in install_out:
        win_version = _get_windows_trellix_version(session)
        send_event(session_id, ip, "installed", _installed_message_with_version(win_version))
        session_results.setdefault(session_id, {})[ip] = {
            "status": "installed",
            "message": "Fresh installation completed from local Downloads path (Windows)",
            "version": win_version,
        }
        return

    if "INSTALL_OK_NOT_RUNNING" in install_out:
        send_event(session_id, ip, "warning", "Installation completed but service is not running", install_out.strip())
        session_results.setdefault(session_id, {})[ip] = {"status": "warning", "message": "Installed but service not running"}
        return

    send_event(session_id, ip, "error", "Installation Failed", install_out.strip() or "Unknown MSI installation error")
    session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": "Windows installation failed"}


def process_server(session_id, server, platform='mxone'):
    """Process a single server - check/install Trellix."""
    config = get_platform_config(platform)
    if platform == 'windows':
        return process_windows_server(session_id, server, config)

    ip = server["ip"]
    admin_username = server["admin_username"]
    admin_password = server["admin_password"]
    root_password = server["root_password"]

    send_event(session_id, ip, "connecting", "Connecting...")

    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

        # Try normal SSH connect
        try:
            ssh.connect(
                hostname=ip,
                username=admin_username,
                password=admin_password,
                timeout=10,
                allow_agent=False,
                look_for_keys=False
            )
        except paramiko.AuthenticationException:
            # Fallback: keyboard-interactive
            try:
                transport = paramiko.Transport((ip, 22))
                transport.start_client()
                transport.auth_interactive(
                    admin_username,
                    lambda title, instructions, prompts: [admin_password] * len(prompts)
                )
                if not transport.is_authenticated():
                    send_event(session_id, ip, "error", "Login Failed",
                              f"Authentication failed for user '{admin_username}'. Check password.")
                    session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": "Authentication failed"}
                    transport.close()
                    return
                ssh._transport = transport
            except Exception as e:
                send_event(session_id, ip, "error", "Login Failed",
                           f"Authentication failed for user '{admin_username}': {str(e)}")
                session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": f"Login failed: {str(e)}"}
                return
        except Exception as e:
            error_msg = str(e)
            if "timed out" in error_msg:
                reason = "Server unreachable (connection timed out). Check IP and network."
            elif "refused" in error_msg:
                reason = "Connection refused. SSH service may not be running on the server."
            elif "No route" in error_msg:
                reason = "No route to host. Server may be offline or IP is incorrect."
            else:
                reason = error_msg
            send_event(session_id, ip, "error", "Connection Failed", reason)
            session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": reason}
            return

        send_event(session_id, ip, "progress", "SSH Login Successful")

        # Open interactive shell
        shell = ssh.invoke_shell()
        time.sleep(1)
        shell.recv(9999)

        # Switch to root
        shell.send("su -\n")
        time.sleep(1)
        output = shell.recv(9999).decode()

        if "Password" in output or "password" in output:
            shell.send(root_password + "\n")
            time.sleep(2)
            output = shell.recv(9999).decode()

        if "#" not in output:
            send_event(session_id, ip, "error", "Root Login Failed",
                       "Could not switch to root. Check root password.")
            session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": "Root login failed"}
            ssh.close()
            return

        send_event(session_id, ip, "progress", "Root Login Successful")

        # Pre-check: Is Trellix already installed?
        send_event(session_id, ip, "progress", "Checking Trellix status...")
        shell.send("systemctl status xagt 2>&1 | head -20\n")
        time.sleep(3)
        status_output = ""
        while shell.recv_ready():
            status_output += shell.recv(65535).decode(errors="replace")
            time.sleep(0.3)

        status_clean = strip_ansi_codes(status_output)

        if "active (running)" in status_clean:
            send_event(session_id, ip, "log", "Trellix (xagt) is already installed and running.")
            # Show key status lines
            for line in status_clean.splitlines():
                line = line.strip()
                if not line:
                    continue
                if any(kw in line for kw in ["Active:", "Main PID:", "Tasks:", "Memory:", "Started xagt"]):
                    send_event(session_id, ip, "log", line)
            linux_version = _get_linux_xagt_version(ssh)
            send_event(session_id, ip, "installed", _installed_message_with_version(linux_version))
            session_results.setdefault(session_id, {})[ip] = {
                "status": "installed",
                "message": "Already installed and running (pre-check)",
                "version": linux_version,
            }
            ssh.close()
            return

        send_event(session_id, ip, "log", "Trellix not running. Proceeding with installation...")

        # Upload fireeye.sh
        local_file = os.path.join(APP_DIR, "fireeye.sh")
        filename = os.path.basename(local_file)
        tmp_path = f"/tmp/{filename}"
        remote_file_path = f"{config['remote_path'].rstrip('/')}/{filename}"

        try:
            sftp = ssh.open_sftp()
            sftp.put(local_file, tmp_path)
            sftp.close()
        except Exception as e:
            send_event(session_id, ip, "error", "Upload Failed", str(e))
            session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": f"Upload failed: {str(e)}"}
            ssh.close()
            return

        send_event(session_id, ip, "progress", "Script Uploaded")

        # Setup file
        shell.settimeout(5)
        for cmd in [
            f"chmod 777 {tmp_path}",
            f"mv {tmp_path} {remote_file_path}",
            f"chmod 777 {remote_file_path}",
        ]:
            shell.send(cmd + "\n")
            time.sleep(1)
            try:
                shell.recv(9999)
            except Exception:
                pass

        # Execute script
        send_event(session_id, ip, "progress", "Running Trellix check...")
        shell.send(f"bash {remote_file_path}\n")
        time.sleep(2)

        full_output = ""
        start = time.time()
        timeout = 180
        pager_quit_sent = False

        while time.time() - start < timeout:
            time.sleep(0.5)
            if not shell.recv_ready():
                continue

            chunk = shell.recv(65535).decode(errors="replace")
            full_output += chunk

            if "Are you sure you want to continue connecting" in chunk:
                shell.send("yes\n")
                time.sleep(1)
                continue
            elif "password:" in chunk.lower() or "password" in chunk.lower():
                # Handle scp/ssh password prompts even when output chunk also contains other text.
                shell.send(config["script_password"] + "\n")
                time.sleep(1)
                continue

            # Detect pager (less/more) and send 'q' to quit it
            if not pager_quit_sent and ("(END)" in chunk or "lines 1-" in chunk or "[7mlines" in chunk):
                shell.send("q")
                time.sleep(1)
                pager_quit_sent = True
                continue

            # Send meaningful lines
            for line in chunk.splitlines():
                stripped = line.strip()
                if not stripped:
                    continue
                
                # Strip ANSI codes and control characters first
                cleaned = strip_ansi_codes(stripped)
                if not cleaned or len(cleaned.strip()) == 0:
                    continue
                cleaned = cleaned.strip()
                
                # Skip terminal control markers and junk
                if cleaned == "------------------------------":
                    continue
                if ":~ #" in cleaned or "bash /root/" in cleaned:
                    continue
                if cleaned == "=" or cleaned == "~" or cleaned == ">":
                    continue
                if cleaned.startswith("lines ") and "END" in cleaned:
                    continue
                if cleaned.startswith("(END)"):
                    continue
                # Skip shell prompts
                if cleaned.startswith("[root@") and cleaned.endswith("#"):
                    continue
                # Skip noisy progress bar lines
                if cleaned.startswith("(100%)") and "#" in cleaned:
                    continue
                
                send_event(session_id, ip, "log", cleaned)

            if "Execution is completed" in chunk:
                break

            # Secondary break: if we see active(running) + Started xagt after pager quit
            if pager_quit_sent and "active (running)" in full_output and "Started xagt" in full_output:
                time.sleep(2)
                # Drain any remaining output
                while shell.recv_ready():
                    extra = shell.recv(65535).decode(errors="replace")
                    full_output += extra
                break

        # Determine final status using both script output and live service check.
        full_output_clean = strip_ansi_codes(full_output)
        final_active = False
        try:
            stdin, stdout, stderr = ssh.exec_command("systemctl is-active xagt 2>/dev/null || true", timeout=8)
            final_active = (stdout.read() or b"").decode(errors="replace").strip().lower() == "active"
        except Exception:
            final_active = False

        if final_active:
            linux_version = _get_linux_xagt_version(ssh)
            send_event(session_id, ip, "installed", _installed_message_with_version(linux_version))
            session_results.setdefault(session_id, {})[ip] = {
                "status": "installed",
                "message": "Installed and service is running",
                "version": linux_version,
            }
        elif "already installed and running" in full_output_clean:
            send_event(session_id, ip, "warning", "Install check inconsistent",
                       "Script reported running, but systemctl check is not active.")
            session_results.setdefault(session_id, {})[ip] = {"status": "warning", "message": "Service not active after script"}
        elif "Execution is completed" in full_output_clean or "Xagent setup is completed" in full_output_clean:
            send_event(session_id, ip, "warning", "Installation completed but service not active",
                       "Script completed, but xagt is not active. Check systemctl status xagt.")
            session_results.setdefault(session_id, {})[ip] = {"status": "warning", "message": "Install completed, service not active"}
        elif "failed" in full_output_clean.lower() or "error" in full_output_clean.lower():
            send_event(session_id, ip, "error", "Installation Failed", "Installer output contains failure markers.")
            session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": "Installation failed"}
        else:
            send_event(session_id, ip, "warning", "Script finished but no confirmation received",
                       "Timed out waiting for 'Execution is completed' message.")
            session_results.setdefault(session_id, {})[ip] = {"status": "warning", "message": "No confirmation received (timeout)"}

        ssh.close()

    except Exception as e:
        send_event(session_id, ip, "error", "Unexpected Error", str(e))
        session_results.setdefault(session_id, {})[ip] = {"status": "error", "message": str(e)}


@app.route('/', strict_slashes=False)
def root_redirect():
    return _serve_index()

@app.route('/trelix', strict_slashes=False)
def trelix_landing():
    return _serve_index()

@app.route('/mivbtrelix', strict_slashes=False)
def mivb_page():
    return _serve_index()

@app.route('/mbgtrelix', strict_slashes=False)
def mbg_page():
    return _serve_index()

@app.route('/mxonetrelix', strict_slashes=False)
def mxone_page():
    return _serve_index()

@app.route('/miv5000trelix', strict_slashes=False)
def miv5000_page():
    return _serve_index()

@app.route('/windowstrelix', strict_slashes=False)
def windows_page():
    return _serve_index()


SMB_SERVER = os.getenv("SMB_SERVER", "10.211.34.118")
SMB_SHARE = os.getenv("SMB_SHARE", "api")
NETWORK_USER = os.getenv("NETWORK_USER", "Administrator")
NETWORK_PASS = os.getenv("NETWORK_PASS", "")
NETWORK_DOMAIN = os.getenv("NETWORK_DOMAIN", "")
SMB_CLIENT_NAME = os.getenv("SMB_CLIENT_NAME", "trellix-manager")
SUPPORTED_EXCEL_EXTENSIONS = ('.xlsx', '.xlsm', '.xltx', '.xltm')

AUTH_ENABLED = str(os.getenv("AUTH_ENABLED", "0")).strip().lower() in ("1", "true", "yes", "on")
PROFILE_TEMPLATES_ENABLED = str(os.getenv("PROFILE_TEMPLATES_ENABLED", "0")).strip().lower() in ("1", "true", "yes", "on")
TOKEN_TTL_SECONDS = int(os.getenv("TOKEN_TTL_SECONDS", "43200"))
_tokens = {}


def _load_users_from_env():
    raw = os.getenv("TRELLIX_USERS_JSON", "").strip()
    if raw:
        try:
            parsed = json.loads(raw)
            users = {}
            for username, val in parsed.items():
                if isinstance(val, dict):
                    users[str(username)] = {
                        "password": str(val.get("password") or ""),
                        "role": str(val.get("role") or "operator").lower(),
                    }
            if users:
                return users
        except Exception as exc:
            log_structured("auth.config_error", error=str(exc))

    # Dev-friendly fallback; override in Docker/K8s with TRELLIX_USERS_JSON.
    return {
        "operator": {"password": os.getenv("OPERATOR_PASSWORD", "operator123"), "role": "operator"},
        "admin": {"password": os.getenv("ADMIN_PASSWORD", "admin123"), "role": "admin"},
    }


APP_USERS = _load_users_from_env()


def _extract_bearer_token(req):
    auth = str(req.headers.get("Authorization") or "")
    if auth.lower().startswith("bearer "):
        return auth[7:].strip()
    qp = str(req.args.get("token") or "").strip()
    if qp:
        return qp
    return ""


def _auth_context(req):
    if not AUTH_ENABLED:
        return {"username": "anonymous", "role": "admin"}
    token = _extract_bearer_token(req)
    if not token:
        return None
    rec = _tokens.get(token)
    if not rec:
        return None
    if time.time() > rec.get("expires_at", 0):
        _tokens.pop(token, None)
        return None
    return rec


def _require_roles(*roles):
    allowed = set(r.lower() for r in roles)

    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            ctx = _auth_context(request)
            if not ctx:
                return jsonify({"error": "Unauthorized"}), 401
            role = str(ctx.get("role") or "").lower()
            if role not in allowed:
                return jsonify({"error": "Forbidden"}), 403
            return fn(*args, **kwargs)

        return wrapper

    return deco


def _append_audit_event(session_id, platform, ip, phase, result, message, details=""):
    if not session_id:
        return
    row = {
        "ts": datetime.utcnow().isoformat() + "Z",
        "session_id": session_id,
        "platform": platform,
        "ip": ip,
        "phase": phase,
        "result": result,
        "message": str(message or ""),
        "details": str(details or ""),
    }
    run_audit_events.setdefault(session_id, []).append(row)


def _audit_file_path(session_id, ext):
    return os.path.join(AUX_DIR, f"run_{session_id}.{ext}")


def _write_audit_exports(session_id):
    events = run_audit_events.get(session_id) or []
    if not events:
        return

    headers = ["ts", "session_id", "platform", "ip", "phase", "result", "message", "details"]
    csv_path = _audit_file_path(session_id, "csv")
    xlsx_path = _audit_file_path(session_id, "xlsx")

    with open(csv_path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in events:
            writer.writerow({k: row.get(k, "") for k in headers})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "audit"
    ws.append(headers)
    for row in events:
        ws.append([row.get(k, "") for k in headers])
    wb.save(xlsx_path)


def _load_profile_templates():
    if not os.path.exists(PROFILE_TEMPLATES_FILE):
        return {}
    try:
        with open(PROFILE_TEMPLATES_FILE, "r", encoding="utf-8") as fh:
            data = json.load(fh)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_profile_templates(data):
    tmp = PROFILE_TEMPLATES_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2)
    os.replace(tmp, PROFILE_TEMPLATES_FILE)


def _validate_package_source(platform, source_path=None, required_files=None):
    config = get_platform_config(platform) or {}
    required_files = required_files or []
    source = str(source_path or config.get("share_dir") or "").strip()
    if platform == "windows" and not source:
        source = str(config.get("installer_unc") or "").strip()

    checks = []
    overall = True
    installer_found = False
    zip_valid = None

    if not source:
        return {
            "ok": False,
            "source": source,
            "checks": [{"name": "path_exists", "ok": False, "message": "No source path configured"}],
            "installer_found": False,
            "zip_valid": None,
        }

    # Validate local path first.
    if os.path.exists(source):
        checks.append({"name": "path_exists", "ok": True, "message": "Source path exists"})
        if os.path.isdir(source):
            all_files = []
            for root, _, files in os.walk(source):
                for f in files:
                    all_files.append(os.path.join(root, f))
            installer_candidates = [f for f in all_files if f.lower().endswith((".msi", ".exe"))]
            installer_found = bool(installer_candidates)
            checks.append({
                "name": "installer_present",
                "ok": installer_found,
                "message": "Installer found" if installer_found else "No MSI/EXE installer found in source",
            })
            zip_candidates = [f for f in all_files if f.lower().endswith(".zip")]
            if zip_candidates:
                try:
                    import zipfile
                    for zf in zip_candidates:
                        with zipfile.ZipFile(zf, "r") as z:
                            names = [n.lower() for n in z.namelist()]
                            if any(n.endswith(".msi") or n.endswith(".exe") for n in names):
                                zip_valid = True
                                break
                    if zip_valid is None:
                        zip_valid = False
                except Exception as exc:
                    zip_valid = False
                    checks.append({"name": "zip_read", "ok": False, "message": f"ZIP validation failed: {exc}"})
            else:
                zip_valid = True

            checks.append({
                "name": "zip_contains_installer",
                "ok": bool(zip_valid),
                "message": "ZIP content valid" if zip_valid else "ZIP does not contain MSI/EXE",
            })

            for rf in required_files:
                found = any(os.path.basename(p).lower() == str(rf).lower() for p in all_files)
                checks.append({
                    "name": f"required:{rf}",
                    "ok": found,
                    "message": f"Required file '{rf}' present" if found else f"Required file '{rf}' missing",
                })
        else:
            lf = source.lower()
            installer_found = lf.endswith((".msi", ".exe"))
            checks.append({
                "name": "installer_present",
                "ok": installer_found,
                "message": "Installer file selected" if installer_found else "Source is not MSI/EXE",
            })
            if lf.endswith(".zip"):
                try:
                    import zipfile
                    with zipfile.ZipFile(source, "r") as z:
                        names = [n.lower() for n in z.namelist()]
                        zip_valid = any(n.endswith(".msi") or n.endswith(".exe") for n in names)
                except Exception as exc:
                    zip_valid = False
                    checks.append({"name": "zip_read", "ok": False, "message": f"ZIP validation failed: {exc}"})
                checks.append({
                    "name": "zip_contains_installer",
                    "ok": bool(zip_valid),
                    "message": "ZIP content valid" if zip_valid else "ZIP does not contain MSI/EXE",
                })
            else:
                zip_valid = True
    else:
        # Best-effort remote SMB validation.
        is_unc = source.startswith("\\\\")
        if is_unc:
            try:
                unc_path = source.lstrip("\\")
                parts = unc_path.split("\\")
                share = parts[1] if len(parts) > 1 else SMB_SHARE
                rel = "/".join(parts[2:]) if len(parts) > 2 else ""
                conn = _smb_conn()
                try:
                    entries = conn.listPath(share, rel.strip("/"))
                    file_names = [e.filename for e in entries if not e.isDirectory]
                    installer_found = any(f.lower().endswith((".msi", ".exe")) for f in file_names)
                    checks.append({"name": "path_exists", "ok": True, "message": "UNC path reachable via SMB"})
                    checks.append({
                        "name": "installer_present",
                        "ok": installer_found,
                        "message": "Installer found" if installer_found else "No MSI/EXE installer found",
                    })
                    zip_candidates = [f for f in file_names if f.lower().endswith(".zip")]
                    zip_valid = True if not zip_candidates else None
                    if zip_candidates:
                        for z in zip_candidates:
                            buf = io.BytesIO()
                            try:
                                conn.retrieveFile(share, f"{rel.strip('/')}/{z}", buf)
                                buf.seek(0)
                                import zipfile
                                with zipfile.ZipFile(buf, "r") as zf:
                                    names = [n.lower() for n in zf.namelist()]
                                    if any(n.endswith(".msi") or n.endswith(".exe") for n in names):
                                        zip_valid = True
                                        break
                            except Exception:
                                zip_valid = False
                        if zip_valid is None:
                            zip_valid = False
                    checks.append({
                        "name": "zip_contains_installer",
                        "ok": bool(zip_valid),
                        "message": "ZIP content valid" if zip_valid else "ZIP does not contain MSI/EXE",
                    })
                finally:
                    conn.close()
            except Exception as exc:
                checks.append({"name": "path_exists", "ok": False, "message": f"UNC path not reachable: {exc}"})
        else:
            checks.append({"name": "path_exists", "ok": False, "message": "Source path does not exist"})

    for c in checks:
        if c.get("ok") is False:
            overall = False

    if platform != "windows":
        fireeye_ok = os.path.exists(os.path.join(APP_DIR, "fireeye.sh"))
        checks.append({
            "name": "required:fireeye.sh",
            "ok": fireeye_ok,
            "message": "fireeye.sh found" if fireeye_ok else "fireeye.sh missing in app directory",
        })
        overall = overall and fireeye_ok

    return {
        "ok": bool(overall),
        "source": source,
        "checks": checks,
        "installer_found": installer_found,
        "zip_valid": zip_valid,
    }


def _is_retryable_timeout_result(result):
    if not isinstance(result, dict):
        return False
    status = str(result.get("status") or "").lower()
    if status != "error":
        return False
    msg = str(result.get("message") or "").lower()
    markers = [
        "timed out",
        "timeout",
        "unreachable",
        "no route",
        "connection dropped",
        "connection reset",
        "connection refused",
    ]
    return any(m in msg for m in markers)


def process_server_with_retry(session_id, server, platform='mxone', max_retries=2):
    ip = server.get("ip", "")
    started = time.time()
    for attempt in range(0, int(max_retries) + 1):
        if attempt > 0:
            send_event(session_id, ip, "warning", f"Retry attempt {attempt}/{max_retries} after network timeout")
        process_server(session_id, server, platform)
        result = session_results.get(session_id, {}).get(ip, {})
        if not _is_retryable_timeout_result(result):
            break
    duration_ms = int((time.time() - started) * 1000)
    final = session_results.get(session_id, {}).get(ip, {})
    log_structured(
        "vm.process.completed",
        vm_ip=ip,
        platform=platform,
        phase="deployment",
        result=str(final.get("status") or "unknown"),
        duration_ms=duration_ms,
    )


def _build_precheck_scorecard(state, reason, credential_valid):
    rs = str(reason or "").lower()
    reachable = not any(k in rs for k in ("unreachable", "not responding", "no route", "timed out", "timeout"))
    platform_match = not ("platform mismatch" in rs)
    trellix_installed = str(state or "") == "installed"
    cred_label = "Valid" if credential_valid is True else ("Invalid" if credential_valid is False else "Unknown")
    return {
        "reachable": "Yes" if reachable else "No",
        "credential_valid": cred_label,
        "platform_match": "Yes" if platform_match else "No",
        "trellix_installed": "Yes" if trellix_installed else "No",
        "ready": "Yes" if (reachable and platform_match and credential_valid is True and not trellix_installed and state == "pending") else "No",
    }


run_audit_events = {}


def _is_supported_excel_file(filename):
    """Return True only for real Excel workbook files (exclude temp/lock artifacts)."""
    name = str(filename or "").strip()
    lower = name.lower()
    if not lower.endswith(SUPPORTED_EXCEL_EXTENSIONS):
        return False
    # Excel lock/temp files appear as "~$<name>.xlsx" while workbook is open.
    if lower.startswith("~$"):
        return False
    return True


def _smb_conn():
    """Return an authenticated SMBConnection to the file server."""
    conn = SMBConnection(NETWORK_USER, NETWORK_PASS, SMB_CLIENT_NAME, SMB_SERVER, use_ntlm_v2=True, is_direct_tcp=True)
    conn.connect(SMB_SERVER, 445)
    return conn


def _normalize_smb_path(path):
    """Normalize SMB paths for pysmb APIs (no leading slash, forward slashes only)."""
    return path.replace('\\', '/').strip('/')


def _candidate_share_dirs(platform):
    """Return possible SMB subpaths for a platform, newest path first."""
    config = get_platform_config(platform) or {}
    configured = config.get("share_dir", "")
    candidates = []
    for raw in (
        configured,
        f"/upgrade_credentials/{platform}",
        f"/trelix_credentails/{platform}",
        f"/trelix_credentials/{platform}",
        f"/trelix/{platform}",
        f"/{platform}",
    ):
        norm = _normalize_smb_path(raw)
        if norm and norm not in candidates:
            candidates.append(norm)
    return candidates


def list_smb_xlsx_files(platform):
    """List all modern Excel files on the network share."""
    if not get_platform_config(platform):
        return [], "Unsupported platform"

    candidates = _candidate_share_dirs(platform)
    try:
        conn = _smb_conn()
        try:
            errors = []
            for share_dir in candidates:
                try:
                    entries = conn.listPath(SMB_SHARE, share_dir)
                    files = [
                        e.filename for e in entries
                        if not e.isDirectory and _is_supported_excel_file(e.filename)
                    ]
                    return files, None
                except Exception as exc:
                    errors.append(f"{share_dir}: {exc}")
        finally:
            conn.close()

        return [], (
            f"Unable to access platform folder in SMB share '{SMB_SHARE}'. "
            f"Tried: {', '.join(candidates)}"
        )
    except Exception as exc:
        return [], str(exc)


def get_smb_file_bytes(platform, filename):
    """Download a file from the network share and return as BytesIO."""
    if not get_platform_config(platform):
        raise ValueError("Unsupported platform")

    candidates = _candidate_share_dirs(platform)
    conn = _smb_conn()
    try:
        for share_dir in candidates:
            buf = io.BytesIO()
            try:
                conn.retrieveFile(SMB_SHARE, f"{share_dir}/{filename}", buf)
                buf.seek(0)
                return buf
            except Exception:
                continue
        raise FileNotFoundError(
            f"File '{filename}' was not found in share '{SMB_SHARE}'. "
            f"Checked: {', '.join(candidates)}"
        )
    finally:
        conn.close()


@app.route('/list-files/<platform>')
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def list_files(platform):
    """Return all Excel files from local storage and network share."""
    if not get_platform_config(platform):
        return jsonify({"error": "Unsupported platform"}), 400

    # By default, keep dropdown clean with network files only.
    include_local = str(request.args.get('include_local', '0')).lower() in ('1', 'true', 'yes')

    results = []
    seen_names = set()

    if include_local:
        # Optional local files from app root + uploads folder.
        for base_dir in (APP_DIR, app.config['UPLOAD_FOLDER']):
            try:
                entries = os.listdir(base_dir)
            except Exception:
                continue

            for f in entries:
                if not _is_supported_excel_file(f):
                    continue
                if not os.path.isfile(os.path.join(base_dir, f)):
                    continue
                name_key = f.lower()
                if name_key in seen_names:
                    continue
                seen_names.add(name_key)
                results.append({"label": f, "value": "local:" + f})

    # Network share files
    network_files, network_error = list_smb_xlsx_files(platform)
    for f in network_files:
        name_key = f.lower()
        if name_key in seen_names:
            continue
        seen_names.add(name_key)
        results.append({"label": f, "value": "network:" + f})

    results.sort(key=lambda x: x["label"].lower())
    return jsonify({"files": results, "network_error": network_error})


@app.route('/use-existing/<platform>', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def use_existing(platform):
    """Run process using an existing xlsx file from the project folder or network share."""
    config = get_platform_config(platform)
    if not config:
        return jsonify({"error": "Unsupported platform"}), 400

    data = request.get_json(silent=True) or {}
    value = data.get('filename', '')
    requested_targets = set(data.get('target_ips') or [])
    precheck_installed_details = data.get('precheck_installed_details') or []
    precheck_by_ip = {}
    for item in precheck_installed_details:
        if not isinstance(item, dict):
            continue
        ip = str(item.get('ip', '')).strip()
        if not ip:
            continue
        precheck_by_ip[ip] = {
            'status': 'installed',
            'message': str(item.get('reason') or 'Already installed and running (precheck)'),
            'version': str(item.get('version') or ''),
            'credential_valid': True,
        }

    servers, err_msg, err_code = _servers_from_existing_file(platform, value)
    if err_msg:
        return jsonify({"error": err_msg}), err_code

    if not servers:
        return jsonify({"error": "No servers found in the Excel file"}), 400

    servers = _dedupe_servers_by_ip(servers)
    servers.sort(key=lambda s: [int(x) for x in s["ip"].split(".")])
    all_servers = list(servers)
    all_server_ips = [s["ip"] for s in servers]

    if requested_targets:
        servers = [server for server in servers if server["ip"] in requested_targets]

    if not servers:
        return jsonify({"error": "No VMs left to deploy after precheck filtering"}), 400

    session_id = str(int(time.time() * 1000))
    event_queues[session_id] = queue.Queue()
    session_meta[session_id] = {"platform": platform}
    run_id = _create_deployment_run(request, platform, value)
    retry_count = int((data.get("retry_count") if isinstance(data.get("retry_count"), int) else 2) or 2)

    def run_all():
        threads = []
        for server in servers:
            t = threading.Thread(target=process_server_with_retry, args=(session_id, server, platform, retry_count))
            t.start()
            threads.append(t)
        for t in threads:
            t.join()

        merged_results = dict(session_results.get(session_id, {}))
        for s in all_servers:
            ip = s.get("ip")
            if not ip or ip in merged_results:
                continue
            if ip in precheck_by_ip:
                merged_results[ip] = precheck_by_ip[ip]

        session_results[session_id] = merged_results

        _save_run_results_to_db(
            run_id=run_id,
            platform=platform,
            servers=all_servers,
            per_ip_results=merged_results,
        )

        send_event(session_id, "__done__", "done", "All servers processed")
        send_summary_email(session_id, platform)

    threading.Thread(target=run_all, daemon=True).start()
    return jsonify({
        "session_id": session_id,
        "servers": [s["ip"] for s in servers],
        "all_servers": all_server_ips,
        "platform_title": config["title"],
    })


@app.route('/precheck-existing/<platform>', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def precheck_existing(platform):
    """Run precheck only and return which VMs already have Trellix installed."""
    config = get_platform_config(platform)
    if not config:
        return jsonify({"error": "Unsupported platform"}), 400

    data = request.get_json(silent=True) or {}
    value = data.get('filename', '')
    servers, err_msg, err_code = _servers_from_existing_file(platform, value)
    if err_msg:
        return jsonify({"error": err_msg}), err_code

    if not servers:
        return jsonify({"error": "No servers found in the Excel file"}), 400

    servers = _dedupe_servers_by_ip(servers)
    servers.sort(key=lambda s: [int(x) for x in s["ip"].split('.')])
    installed, installed_details, pending, errors, precheck_details = _run_precheck(platform, servers)

    precheck_results = _precheck_details_to_results(precheck_details)
    precheck_run_id = _create_deployment_run(request, platform, f"precheck:{value}")
    # Save precheck results in background so UI is not blocked at "Checking...".
    threading.Thread(
        target=_save_run_results_to_db,
        kwargs={
            "run_id": precheck_run_id,
            "platform": platform,
            "servers": servers,
            "per_ip_results": precheck_results,
        },
        daemon=True,
    ).start()

    return jsonify({
        "platform_title": config["title"],
        "all_servers": [s["ip"] for s in servers],
        "installed": installed,
        "installed_details": installed_details,
        "to_deploy": [s["ip"] for s in pending],
        "errors": errors,
        "precheck_details": precheck_details,
    })


def _server_from_manual_input(platform, data):
    """Build one server from manual form input with platform-specific requirements."""
    ip = str((data or {}).get('ip') or '').strip()
    username = str((data or {}).get('admin_username') or '').strip()
    password = str((data or {}).get('admin_password') or '').strip()
    root_password = str((data or {}).get('root_password') or '').strip()

    if not ip or not username or not password:
        return None, "IP, Putty Username and Putty Password are required", 400

    if platform == 'mxone' and not root_password:
        return None, "MxOne requires su - (root) password", 400

    cfg = get_platform_config(platform) or {}
    configured_root = str(cfg.get('root_password', '') or '').strip()
    if platform != 'mxone' and configured_root:
        # Non-mxone Linux platforms use fixed root/script password policy.
        root_password = configured_root
    elif not root_password:
        root_password = configured_root

    return {
        "ip": ip,
        "admin_username": username,
        "admin_password": password,
        "root_password": root_password,
    }, None, None


@app.route('/precheck-single/<platform>', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def precheck_single(platform):
    """Run precheck for one manually entered server."""
    config = get_platform_config(platform)
    if not config:
        return jsonify({"error": "Unsupported platform"}), 400

    data = request.get_json(silent=True) or {}
    server, err_msg, err_code = _server_from_manual_input(platform, data)
    if err_msg:
        return jsonify({"error": err_msg}), err_code

    installed, installed_details, pending, errors, precheck_details = _run_precheck(platform, [server])

    precheck_results = _precheck_details_to_results(precheck_details)
    precheck_run_id = _create_deployment_run(request, platform, 'precheck:manual:single')
    # Save precheck results in background so UI is not blocked at "Checking...".
    threading.Thread(
        target=_save_run_results_to_db,
        kwargs={
            "run_id": precheck_run_id,
            "platform": platform,
            "servers": [server],
            "per_ip_results": precheck_results,
        },
        daemon=True,
    ).start()

    return jsonify({
        "platform_title": config["title"],
        "all_servers": [server["ip"]],
        "installed": installed,
        "installed_details": installed_details,
        "to_deploy": [s["ip"] for s in pending],
        "errors": errors,
        "precheck_details": precheck_details,
    })


@app.route('/use-single/<platform>', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def use_single(platform):
    """Deploy Trellix for one manually entered server."""
    config = get_platform_config(platform)
    if not config:
        return jsonify({"error": "Unsupported platform"}), 400

    data = request.get_json(silent=True) or {}
    requested_targets = set(data.get('target_ips') or [])
    precheck_installed_details = data.get('precheck_installed_details') or []

    precheck_by_ip = {}
    for item in precheck_installed_details:
        if not isinstance(item, dict):
            continue
        ip = str(item.get('ip', '')).strip()
        if not ip:
            continue
        precheck_by_ip[ip] = {
            'status': 'installed',
            'message': str(item.get('reason') or 'Already installed and running (precheck)'),
            'version': str(item.get('version') or ''),
            'credential_valid': True,
        }

    server, err_msg, err_code = _server_from_manual_input(platform, data)
    if err_msg:
        return jsonify({"error": err_msg}), err_code

    all_servers = [server]
    servers = list(all_servers)
    if requested_targets:
        servers = [s for s in servers if s["ip"] in requested_targets]

    if not servers:
        return jsonify({"error": "No VMs left to deploy after precheck filtering"}), 400

    session_id = str(int(time.time() * 1000))
    event_queues[session_id] = queue.Queue()
    session_meta[session_id] = {"platform": platform}
    run_id = _create_deployment_run(request, platform, 'manual:single')
    retry_count = int((data.get("retry_count") if isinstance(data.get("retry_count"), int) else 2) or 2)

    def run_all():
        threads = []
        for s in servers:
            t = threading.Thread(target=process_server_with_retry, args=(session_id, s, platform, retry_count))
            t.start()
            threads.append(t)
        for t in threads:
            t.join()

        merged_results = dict(session_results.get(session_id, {}))
        for s in all_servers:
            ip = s.get("ip")
            if not ip or ip in merged_results:
                continue
            if ip in precheck_by_ip:
                merged_results[ip] = precheck_by_ip[ip]

        session_results[session_id] = merged_results
        _save_run_results_to_db(
            run_id=run_id,
            platform=platform,
            servers=all_servers,
            per_ip_results=merged_results,
        )

        send_event(session_id, "__done__", "done", "All servers processed")
        send_summary_email(session_id, platform)

    threading.Thread(target=run_all, daemon=True).start()
    return jsonify({
        "session_id": session_id,
        "servers": [s["ip"] for s in servers],
        "all_servers": [server["ip"]],
        "platform_title": config["title"],
    })


def _servers_from_existing_file(platform, value):
    """Load server rows from local or network xlsx selection.

    Returns: (servers, err_msg, err_code)
    """
    if not value:
        return None, "No file selected", 400

    if value.startswith('network:'):
        filename = os.path.basename(value[len('network:'):])
        try:
            file_obj = get_smb_file_bytes(platform, filename)
        except Exception as e:
            return None, f"Could not retrieve file from network share: {str(e)}", 404
        try:
            return load_servers_from_file(file_obj, platform), None, None
        except Exception as e:
            return None, f"Failed to read Excel file: {str(e)}", 400

    filename = os.path.basename(value.replace('local:', '', 1))
    local_candidates = [
        os.path.join(APP_DIR, filename),
        os.path.join(app.config['UPLOAD_FOLDER'], filename),
    ]
    filepath = next((p for p in local_candidates if os.path.exists(p)), None)
    if not filepath:
        return None, "File not found", 404

    try:
        return load_servers_from_file(filepath, platform), None, None
    except Exception as e:
        return None, f"Failed to read Excel file: {str(e)}", 400


def _precheck_one_server(platform, server):
    """Return (state, reason, version, credential_valid) for precheck.

    credential_valid values:
    - True: Authentication succeeded.
    - False: Authentication explicitly failed.
    - None: Could not determine due to connectivity/other runtime failure.
    """
    ip = server.get("ip", "")
    username = server.get("admin_username", "")
    password = server.get("admin_password", "")

    # Step 1: Ping is advisory only; many environments block ICMP while SSH/WinRM still works.
    ping_ok = _is_ip_pingable(ip, timeout=2)

    if platform == 'windows':
        cfg = get_platform_config('windows') or {}
        winrm_port = int(cfg.get('winrm_port', 5985))
        endpoint = f"http://{ip}:{winrm_port}/wsman"
        transport = cfg.get("winrm_transport", "ntlm")

        # Windows flow should use WinRM reachability, not SSH/PuTTY port checks.
        if not _is_system_reachable(ip, port=winrm_port, timeout=5):
            if not ping_ok:
                return 'error', (
                    f'Precheck failed: IP {ip} not reachable by ping and WinRM not reachable on {winrm_port}'
                ), '', None
            if _is_system_reachable(ip, port=3389, timeout=3):
                return 'error', (
                    f'Precheck failed: WinRM not reachable on {ip}:{winrm_port}. '
                    'RDP appears reachable; enable WinRM (winrm quickconfig) and open firewall port 5985.'
                ), '', None
            return 'error', (
                f'Precheck failed: System unreachable for WinRM (IP {ip}, port {winrm_port} not responding)'
            ), '', None

        try:
            sess = winrm.Session(endpoint, auth=(username, password), transport=transport)
            cmd = r"""
$svc = Get-Service -Name xagt -ErrorAction SilentlyContinue
if (-not $svc) { $svc = Get-Service -Name masvc -ErrorAction SilentlyContinue }
if ($svc -and $svc.Status -eq 'Running') { Write-Output 'INSTALLED_RUNNING' }
elseif ($svc) { Write-Output 'INSTALLED_STOPPED' }
else { Write-Output 'NOT_INSTALLED' }
"""
            res = sess.run_ps(cmd)
            out = ((res.std_out or b"") + (res.std_err or b"")).decode(errors="replace").lower()
            if 'installed_running' in out:
                win_version = _get_windows_trellix_version(sess)
                return 'installed', 'Precheck: Trellix already running (credentials valid)', win_version or '', True
            if 'installed_stopped' in out:
                win_version = _get_windows_trellix_version(sess)
                return 'installed', 'Precheck: Trellix installed but service is not running (credentials valid)', win_version or '', True
            return 'pending', 'Precheck: Trellix not running (credentials valid)', '', True
        except Exception as exc:
            msg = str(exc).lower()
            if any(k in msg for k in ('401', 'unauthorized', 'auth', 'access denied', 'invalid')):
                return 'error', 'Precheck failed: credentials invalid', '', False
            return 'error', f'Precheck failed: {exc}', '', None

    # Step 1 (Linux platforms): Check if system is reachable over SSH
    if not _is_system_reachable(ip, port=22, timeout=5):
        if not ping_ok:
            return 'error', f'Precheck failed: IP {ip} not reachable by ping and SSH port 22 not responding', '', None
        return 'error', f'Precheck failed: System unreachable (IP {ip} on SSH port 22 not responding)', '', None

    # Linux-style precheck over SSH for mxone/mivb/miv5000/mbg
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        try:
            ssh.connect(
                hostname=ip,
                username=username,
                password=password,
                timeout=10,
                auth_timeout=10,
                banner_timeout=10,
                allow_agent=False,
                look_for_keys=False,
            )
        except paramiko.AuthenticationException:
            # Fallback for servers requiring keyboard-interactive authentication.
            transport = paramiko.Transport((ip, 22))
            transport.banner_timeout = 10
            transport.auth_timeout = 10
            transport.start_client(timeout=10)
            transport.auth_interactive(
                username,
                lambda title, instructions, prompts: [password] * len(prompts)
            )
            if not transport.is_authenticated():
                transport.close()
                return 'error', 'Precheck failed: credentials invalid', '', False
            ssh._transport = transport

        identity_ok, identity_reason = _verify_linux_platform_identity(ssh, platform)
        if not identity_ok:
            return 'error', f'Precheck failed: {identity_reason}', '', True

        stdin, stdout, stderr = ssh.exec_command("systemctl is-active xagt 2>/dev/null || true", timeout=8)
        out = (stdout.read() or b"").decode(errors='replace').strip().lower()

        version = ''
        rpm_installed = False
        try:
            stdin, stdout, stderr = ssh.exec_command("rpm -q xagt 2>&1 || true", timeout=8)
            rpm_out = (stdout.read() or b"").decode(errors='replace').strip()
            if rpm_out and 'not installed' not in rpm_out.lower() and 'command not found' not in rpm_out.lower():
                rpm_installed = True
                version = rpm_out.splitlines()[-1].strip()
        except Exception:
            rpm_installed = False
            version = ''

        ssh.close()
        if out == 'active':
            return 'installed', 'Precheck: Trellix already running', version, True
        if rpm_installed:
            return 'installed', 'Precheck: Trellix installed but service is not active', version, True
        return 'pending', 'Precheck: Trellix not active', '', True
    except paramiko.AuthenticationException:
        return 'error', 'Precheck failed: credentials invalid', '', False
    except Exception as exc:
        return 'error', f'Precheck failed: {exc}', '', None
    finally:
        try:
            ssh.close()
        except Exception:
            pass


def _run_precheck(platform, servers):
    installed = []
    installed_details = []
    pending = []
    errors = []
    precheck_details = []

    max_workers = min(30, max(1, len(servers)))
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
        fut_map = {ex.submit(_precheck_one_server, platform, s): s for s in servers}
        for fut in concurrent.futures.as_completed(fut_map):
            server = fut_map[fut]
            ip = server.get('ip')
            try:
                state, reason, version, credential_valid = fut.result()
            except Exception as exc:
                state, reason, version, credential_valid = 'error', f'Precheck failed: {exc}', '', None

            precheck_details.append({
                'ip': ip,
                'credential_valid': credential_valid,
                'installed': state == 'installed',
                'version': version or '',
                'state': state,
                'reason': reason,
                'scorecard': _build_precheck_scorecard(state, reason, credential_valid),
            })

            if state == 'installed':
                installed.append(ip)
                installed_details.append({
                    'ip': ip,
                    'reason': reason,
                    'version': version or ''
                })
            elif state == 'pending':
                pending.append(server)
            else:
                errors.append({'ip': ip, 'reason': reason})

    installed.sort(key=lambda x: [int(p) for p in x.split('.')])
    installed_details.sort(key=lambda x: [int(p) for p in x['ip'].split('.')])
    precheck_details.sort(key=lambda x: [int(p) for p in x['ip'].split('.')])
    pending.sort(key=lambda s: [int(p) for p in s['ip'].split('.')])
    return installed, installed_details, pending, errors, precheck_details


@app.route('/preview-existing/<platform>', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def preview_existing(platform):
    """Return server count from selected file without starting deployment."""
    config = get_platform_config(platform)
    if not config:
        return jsonify({"error": "Unsupported platform"}), 400

    data = request.get_json(silent=True) or {}
    value = data.get('filename', '')
    servers, err_msg, err_code = _servers_from_existing_file(platform, value)
    if err_msg:
        return jsonify({"error": err_msg}), err_code

    servers = _dedupe_servers_by_ip(servers)

    return jsonify({
        "count": len(servers or []),
        "servers": [s.get("ip") for s in (servers or []) if s.get("ip")],
        "platform_title": config["title"],
    })


@app.route('/health', methods=['GET'])
@limiter.exempt
def health():
    return jsonify({"status": "ok", "service": "trelix-manager"})


@app.route('/ready', methods=['GET'])
@limiter.exempt
def ready():
    db_ok = True
    db_error = ""
    if _db_enabled():
        conn = None
        try:
            conn = _get_db_conn()
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
        except Exception as exc:
            db_ok = False
            db_error = str(exc)
        finally:
            if conn:
                conn.close()
    payload = {
        "ready": bool(db_ok),
        "db": "ok" if db_ok else "error",
    }
    if db_error:
        payload["error"] = db_error
    return jsonify(payload), (200 if db_ok else 503)


@app.route('/auth/login', methods=['POST'])
@limiter.limit("20 per minute")
def auth_login():
    if not AUTH_ENABLED:
        return jsonify({"token": "disabled", "role": "admin", "username": "anonymous"})
    data = request.get_json(silent=True) or {}
    username = str(data.get("username") or "").strip()
    password = str(data.get("password") or "")
    user = APP_USERS.get(username)
    if not user or str(user.get("password") or "") != password:
        return jsonify({"error": "Invalid credentials"}), 401
    token = secrets.token_urlsafe(32)
    role = str(user.get("role") or "operator").lower()
    _tokens[token] = {
        "username": username,
        "role": role,
        "issued_at": time.time(),
        "expires_at": time.time() + TOKEN_TTL_SECONDS,
    }
    return jsonify({"token": token, "role": role, "username": username, "expires_in": TOKEN_TTL_SECONDS})


@app.route('/auth/me', methods=['GET'])
@limiter.exempt
def auth_me():
    ctx = _auth_context(request)
    if not ctx:
        return jsonify({"error": "Unauthorized"}), 401
    return jsonify({"username": ctx.get("username"), "role": ctx.get("role")})


@app.route('/profiles/<platform>', methods=['GET'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def list_profiles(platform):
    if not PROFILE_TEMPLATES_ENABLED:
        return jsonify({"error": "Reusable profiles are disabled"}), 404
    if not get_platform_config(platform):
        return jsonify({"error": "Unsupported platform"}), 400
    data = _load_profile_templates()
    return jsonify({"profiles": data.get(platform, [])})


@app.route('/profiles/<platform>', methods=['POST'])
@limiter.limit("50 per minute")
@_require_roles("admin")
def save_profile(platform):
    if not PROFILE_TEMPLATES_ENABLED:
        return jsonify({"error": "Reusable profiles are disabled"}), 404
    if not get_platform_config(platform):
        return jsonify({"error": "Unsupported platform"}), 400
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name") or "").strip()
    template = payload.get("template") or {}
    if not name:
        return jsonify({"error": "Profile name is required"}), 400
    if not isinstance(template, dict):
        return jsonify({"error": "Template must be an object"}), 400

    data = _load_profile_templates()
    bucket = data.setdefault(platform, [])
    existing = next((p for p in bucket if str(p.get("name") or "").lower() == name.lower()), None)
    item = {
        "name": name,
        "template": {
            "share_path": str(template.get("share_path") or ""),
            "installer_type": str(template.get("installer_type") or "auto"),
            "silent_args": str(template.get("silent_args") or ""),
            "ports": template.get("ports") if isinstance(template.get("ports"), list) else [],
        },
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }
    if existing:
        existing.update(item)
    else:
        bucket.append(item)
    _save_profile_templates(data)
    return jsonify({"ok": True, "profile": item})


@app.route('/profiles/<platform>/<name>', methods=['DELETE'])
@limiter.limit("50 per minute")
@_require_roles("admin")
def delete_profile(platform, name):
    if not PROFILE_TEMPLATES_ENABLED:
        return jsonify({"error": "Reusable profiles are disabled"}), 404
    if not get_platform_config(platform):
        return jsonify({"error": "Unsupported platform"}), 400
    data = _load_profile_templates()
    bucket = data.get(platform, [])
    keep = [p for p in bucket if str(p.get("name") or "").lower() != str(name).lower()]
    data[platform] = keep
    _save_profile_templates(data)
    return jsonify({"ok": True})


@app.route('/validate-package-source/<platform>', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def validate_package_source(platform):
    if not ENABLE_RESULT_EXPORTS:
        return jsonify({"error": "Package source validation disabled"}), 404
    if not get_platform_config(platform):
        return jsonify({"error": "Unsupported platform"}), 400
    payload = request.get_json(silent=True) or {}
    source_path = payload.get("source_path")
    required_files = payload.get("required_files") or []
    result = _validate_package_source(platform, source_path=source_path, required_files=required_files)
    return jsonify(result), (200 if result.get("ok") else 400)


@app.route('/runs/<session_id>/export/<fmt>', methods=['GET'])
@limiter.limit("60 per minute")
@_require_roles("operator", "admin")
def export_run_audit(session_id, fmt):
    if not ENABLE_RESULT_EXPORTS:
        return jsonify({"error": "Run export disabled"}), 404
    if fmt not in ("csv", "xlsx"):
        return jsonify({"error": "Unsupported format"}), 400
    _write_audit_exports(session_id)
    path = _audit_file_path(session_id, fmt)
    if not os.path.exists(path):
        return jsonify({"error": "No run audit found"}), 404
    download_name = f"trelix_run_{session_id}.{fmt}"
    mime = "text/csv" if fmt == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return send_file(path, as_attachment=True, download_name=download_name, mimetype=mime)


@app.route('/upload-file', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def upload_file_only():
    """Upload an Excel file into uploads folder and return a selectable value."""
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if not _is_supported_excel_file(file.filename):
        return jsonify({"error": "Only modern Excel files (.xlsx/.xlsm/.xltx/.xltm) are supported"}), 400

    safe_name = secure_filename(file.filename)
    if not safe_name:
        return jsonify({"error": "Invalid filename"}), 400

    base, ext = os.path.splitext(safe_name)
    target_name = safe_name
    target_path = os.path.join(app.config['UPLOAD_FOLDER'], target_name)
    if os.path.exists(target_path):
        target_name = f"{base}_{int(time.time())}{ext}"
        target_path = os.path.join(app.config['UPLOAD_FOLDER'], target_name)

    file.save(target_path)

    log_structured(
        "upload.received",
        vm_ip="",
        platform="n/a",
        phase="upload",
        result="ok",
        duration_ms=0,
        file=target_name,
    )

    return jsonify({
        "label": target_name,
        "value": f"local:{target_name}",
        "filename": target_name,
    })

@app.route('/upload', methods=['POST'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def upload():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if not _is_supported_excel_file(file.filename):
        return jsonify({"error": "Only modern Excel files (.xlsx/.xlsm/.xltx/.xltm) are supported"}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(filepath)

    try:
        servers = load_servers_from_file(filepath, 'mxone')
    except Exception as e:
        return jsonify({"error": f"Failed to read Excel file: {str(e)}"}), 400

    if not servers:
        return jsonify({"error": "No servers found in the Excel file"}), 400

    # Sort by IP ascending
    servers.sort(key=lambda s: [int(x) for x in s["ip"].split(".")])

    # Create session
    session_id = str(int(time.time() * 1000))
    event_queues[session_id] = queue.Queue()

    # Start processing in background
    def run_all():
        threads = []
        for server in servers:
            t = threading.Thread(target=process_server, args=(session_id, server, 'mxone'))
            t.start()
            threads.append(t)
        for t in threads:
            t.join()
        send_event(session_id, "__done__", "done", "All servers processed")
        send_summary_email(session_id, 'mxone')

    threading.Thread(target=run_all, daemon=True).start()

    return jsonify({
        "session_id": session_id,
        "servers": [s["ip"] for s in servers]
    })


@app.route('/stream/<session_id>')
@limiter.exempt
@_require_roles("operator", "admin")
def stream(session_id):
    def generate():
        if session_id not in event_queues:
            return
        q = event_queues[session_id]
        while True:
            try:
                event = q.get(timeout=300)
                yield f"data: {json.dumps(event)}\n\n"
                if event["status"] == "done":
                    break
            except queue.Empty:
                break
        # Cleanup
        del event_queues[session_id]
        session_meta.pop(session_id, None)

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/runs/history', methods=['GET'])
@limiter.limit("100 per minute")
@_require_roles("operator", "admin")
def runs_history():
    limit = int(request.args.get("limit", "50"))
    limit = max(1, min(200, limit))

    if not _db_enabled():
        sessions = []
        for sid, events in run_audit_events.items():
            platform_name = (events[0].get("platform") if events else "") or ""
            sessions.append({
                "session_id": sid,
                "platform": platform_name,
                "events": len(events),
                "started_at": events[0].get("ts") if events else "",
                "completed": any(str(e.get("result")) == "done" for e in events),
            })
        sessions.sort(key=lambda x: x.get("started_at", ""), reverse=True)
        return jsonify({"runs": sessions[:limit]})

    conn = None
    try:
        conn = _get_db_conn()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT r.id, r.vm_type, r.source_file, r.started_at, r.completed_at,
                       COUNT(dr.id) AS total_results,
                       SUM(CASE WHEN dr.status = 'installed' THEN 1 ELSE 0 END) AS installed_count,
                       SUM(CASE WHEN dr.status IN ('error', 'warning') THEN 1 ELSE 0 END) AS failed_count
                FROM deployment_runs r
                LEFT JOIN deployment_results dr ON dr.run_id = r.id
                GROUP BY r.id, r.vm_type, r.source_file, r.started_at, r.completed_at
                ORDER BY r.started_at DESC
                LIMIT %s
                """,
                (limit,),
            )
            rows = cur.fetchall() or []
            runs = []
            for row in rows:
                runs.append({
                    "run_id": row.get("id"),
                    "platform": row.get("vm_type"),
                    "source_file": row.get("source_file"),
                    "started_at": row.get("started_at"),
                    "completed_at": row.get("completed_at"),
                    "total": int(row.get("total_results") or 0),
                    "installed": int(row.get("installed_count") or 0),
                    "failed": int(row.get("failed_count") or 0),
                })
            return jsonify({"runs": runs})
    except Exception as exc:
        return jsonify({"error": f"Failed to load run history: {exc}"}), 500
    finally:
        if conn:
            conn.close()


if __name__ == '__main__':
    host = os.getenv('TRELLIX_BIND_HOST', '0.0.0.0').strip() or '0.0.0.0'
    port = int(os.getenv('TRELLIX_PORT', '5000'))
    debug_mode = str(os.getenv('TRELLIX_DEBUG', '0')).strip().lower() in ('1', 'true', 'yes', 'on')
    if getattr(sys, 'frozen', False):
        # Run as desktop app with native window (no browser, no CMD)
        import webview

        def start_flask():
            app.run(debug=False, host=host, port=port, threaded=True, use_reloader=False)

        # Start Flask server in a background thread
        flask_thread = threading.Thread(target=start_flask, daemon=True)
        flask_thread.start()

        # Wait for Flask to be ready
        import urllib.request
        for _ in range(30):
            try:
                urllib.request.urlopen(f'http://127.0.0.1:{port}/trelix')
                break
            except Exception:
                time.sleep(0.2)

        # Create native window
        webview.create_window('Trellix Manager', f'http://127.0.0.1:{port}/trelix',
                              width=1280, height=800, resizable=True)
        webview.start()
    else:
        # Development mode - run normally with browser
        import webbrowser
        if debug_mode:
            threading.Timer(1.5, lambda: webbrowser.open(f'http://localhost:{port}/trelix')).start()
        app.run(debug=debug_mode, host=host, port=port, threaded=True, use_reloader=debug_mode)
