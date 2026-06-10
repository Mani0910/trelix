import argparse
import concurrent.futures
import getpass
import http.server
import ipaddress
import json
import os
import socket
import socketserver
import tempfile
import threading
from pathlib import Path

import openpyxl
import paramiko
import winrm


def _norm(s: str) -> str:
    return (s or "").strip().lower().replace("_", " ")


def read_vm_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [
        _norm(str(cell.value) if cell.value is not None else "")
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    idx_ip = None
    idx_user = None
    idx_pass = None

    for i, h in enumerate(headers):
        if h in {"server ip", "ip", "server", "host", "hostname"}:
            idx_ip = i
        elif h in {"username", "user", "admin username", "windows username"}:
            idx_user = i
        elif h in {"password", "pass", "admin password", "windows password"}:
            idx_pass = i

    rows = []
    has_header = idx_ip is not None and idx_user is not None and idx_pass is not None

    if has_header:
        for row in ws.iter_rows(min_row=2, values_only=True):
            ip = str(row[idx_ip]).strip() if row[idx_ip] else ""
            user = str(row[idx_user]).strip() if row[idx_user] else ""
            pwd = str(row[idx_pass]).strip() if row[idx_pass] else ""
            if ip and user and pwd:
                rows.append({"ip": ip, "username": user, "password": pwd})
    else:
        # Headerless format: A=IP, B=username, C=password
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            ip = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            user = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            pwd = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            if ip and user and pwd:
                rows.append({"ip": ip, "username": user, "password": pwd})

    return rows


def build_vm_rows_from_args(args):
    """Build target rows from either XLSX or direct --ip/--username/--password args."""
    if args.xlsx:
        return read_vm_rows(args.xlsx)

    if args.ip and args.username and args.password:
        return [{"ip": args.ip, "username": args.username, "password": args.password}]

    raise RuntimeError(
        "Provide either --xlsx <file> OR --ip <ip> --username <user> --password <pass>"
    )


def download_zip_from_linux(host: str, user: str, password: str, remote_zip_path: str, local_zip_path: str):
    transport = paramiko.Transport((host, 22))
    try:
        transport.connect(username=user, password=password)
        sftp = paramiko.SFTPClient.from_transport(transport)
        try:
            sftp.get(remote_zip_path, local_zip_path)
        finally:
            sftp.close()
    finally:
        transport.close()


def pick_local_ip():
    # Best-effort IPv4 address used by VMs to reach this machine.
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        return s.getsockname()[0]
    finally:
        s.close()


class QuietHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format, *args):
        return


def start_http_server(serve_dir: str, port: int):
    class _Handler(QuietHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=serve_dir, **kwargs)

    server = socketserver.ThreadingTCPServer(("0.0.0.0", port), _Handler)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server


def transfer_and_unzip_on_vm(vm, package_url: str, package_name: str, transport: str, port: int):
    endpoint = f"http://{vm['ip']}:{port}/wsman"
    sess = winrm.Session(endpoint, auth=(vm["username"], vm["password"]), transport=transport)

    ps = f"""
$downloads = Join-Path $env:USERPROFILE 'Downloads'
New-Item -Path $downloads -ItemType Directory -Force | Out-Null
$zipPath = Join-Path $downloads '{package_name}'
$extractDir = Join-Path $downloads ([IO.Path]::GetFileNameWithoutExtension('{package_name}'))

Invoke-WebRequest -UseBasicParsing -Uri '{package_url}' -OutFile $zipPath
if (-not (Test-Path $zipPath)) {{
  Write-Output 'ZIP_COPY_FAILED'
  exit 1
}}

Expand-Archive -Path $zipPath -DestinationPath $extractDir -Force
if (-not (Test-Path $extractDir)) {{
  Write-Output 'UNZIP_FAILED'
  exit 1
}}

Write-Output ('ZIP_PATH=' + $zipPath)
Write-Output ('EXTRACT_DIR=' + $extractDir)
Write-Output 'DONE'
"""

    r = sess.run_ps(ps)
    out = ((r.std_out or b"") + (r.std_err or b"")).decode(errors="replace")

    if "DONE" in out:
        return True, out.strip()
    return False, out.strip() or "Unknown error"


def _parse_marker(output: str, key: str) -> str:
        prefix = f"{key}="
        for line in output.splitlines():
                if line.startswith(prefix):
                        return line[len(prefix):].strip()
        return ""


def transfer_unzip_install_and_check(vm, package_url: str, package_name: str, transport: str, port: int):
        endpoint = f"http://{vm['ip']}:{port}/wsman"
        sess = winrm.Session(endpoint, auth=(vm["username"], vm["password"]), transport=transport)

        ps = f"""
$downloads = Join-Path $env:USERPROFILE 'Downloads'
New-Item -Path $downloads -ItemType Directory -Force | Out-Null

$zipPath = Join-Path $downloads '{package_name}'
$extractDir = Join-Path $downloads ([IO.Path]::GetFileNameWithoutExtension('{package_name}'))

Invoke-WebRequest -UseBasicParsing -Uri '{package_url}' -OutFile $zipPath
if (-not (Test-Path $zipPath)) {{
    Write-Output 'RESULT=ZIP_COPY_FAILED'
    exit 1
}}

Expand-Archive -Path $zipPath -DestinationPath $extractDir -Force
if (-not (Test-Path $extractDir)) {{
    Write-Output 'RESULT=UNZIP_FAILED'
    exit 1
}}

$msi = Get-ChildItem -Path $extractDir -Recurse -File -Filter *.msi -ErrorAction SilentlyContinue |
    Select-Object -First 1

if (-not $msi) {{
    Write-Output 'RESULT=MSI_NOT_FOUND'
    Write-Output ('ZIP_PATH=' + $zipPath)
    Write-Output ('EXTRACT_DIR=' + $extractDir)
    exit 1
}}

$proc = Start-Process -FilePath msiexec.exe -ArgumentList "/i `"$($msi.FullName)`" /qn /norestart" -Wait -PassThru

# Try to start known Trellix services if present.
$svc = Get-Service -Name xagt -ErrorAction SilentlyContinue
if (-not $svc) {{ $svc = Get-Service -Name masvc -ErrorAction SilentlyContinue }}
if ($svc) {{
    Start-Service -Name $svc.Name -ErrorAction SilentlyContinue
    $svc = Get-Service -Name $svc.Name -ErrorAction SilentlyContinue
}}

# Best-effort version lookup.
$ver = ''
$paths = @(
    'C:\Program Files\McAfee\Agent\x64\xagt.exe',
    'C:\Program Files\McAfee\Agent\xagt.exe'
)
foreach ($p in $paths) {{
    if (Test-Path $p) {{
        $ver = (Get-Item $p).VersionInfo.ProductVersion
        if ($ver) {{ break }}
    }}
}}

Write-Output ('ZIP_PATH=' + $zipPath)
Write-Output ('EXTRACT_DIR=' + $extractDir)
Write-Output ('MSI_PATH=' + $msi.FullName)
Write-Output ('INSTALL_EXIT=' + $proc.ExitCode)
Write-Output ('SERVICE_NAME=' + $(if ($svc) {{ $svc.Name }} else {{ '' }}))
Write-Output ('SERVICE_STATUS=' + $(if ($svc) {{ $svc.Status }} else {{ 'NOT_FOUND' }}))
Write-Output ('VERSION=' + $ver)

if ($proc.ExitCode -ne 0) {{
    Write-Output 'RESULT=INSTALL_COMMAND_FAILED'
    exit 1
}}

if ($svc -and $svc.Status -eq 'Running') {{
    Write-Output 'RESULT=INSTALLED_RUNNING'
}} elseif ($svc) {{
    Write-Output 'RESULT=INSTALLED_SERVICE_NOT_RUNNING'
}} else {{
    Write-Output 'RESULT=INSTALLED_SERVICE_NOT_FOUND'
}}
"""

        r = sess.run_ps(ps)
        out = ((r.std_out or b"") + (r.std_err or b"")).decode(errors="replace").strip()

        result = _parse_marker(out, "RESULT")
        info = {
                "zip_path": _parse_marker(out, "ZIP_PATH"),
                "extract_dir": _parse_marker(out, "EXTRACT_DIR"),
                "msi_path": _parse_marker(out, "MSI_PATH"),
                "install_exit": _parse_marker(out, "INSTALL_EXIT"),
                "service_name": _parse_marker(out, "SERVICE_NAME"),
                "service_status": _parse_marker(out, "SERVICE_STATUS"),
                "version": _parse_marker(out, "VERSION"),
                "result": result,
                "raw": out,
        }

        success_states = {
                "INSTALLED_RUNNING",
                "INSTALLED_SERVICE_NOT_RUNNING",
                "INSTALLED_SERVICE_NOT_FOUND",
        }
        success = result in success_states
        return success, info


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Transfer ZIP from Linux to Windows VM(s), unzip, install MSI, "
            "and report install status via WinRM"
        )
    )
    parser.add_argument("--xlsx", help="Path to XLSX with server ip, username, password")
    parser.add_argument("--ip", help="Single Windows VM IP")
    parser.add_argument("--username", help="Single Windows VM username")
    parser.add_argument("--password", help="Single Windows VM password")
    parser.add_argument("--linux-host", default="10.211.27.74")
    parser.add_argument("--linux-user", default="root")
    parser.add_argument("--linux-pass", default=None, help="Linux password; if omitted prompt securely")
    parser.add_argument("--linux-zip-path", default="/home/Fireeye/IMAGE_HX_AGENT_WIN_36.30.37.zip")
    parser.add_argument("--http-host", default=None, help="IP/host reachable by Windows VMs for package download")
    parser.add_argument("--http-port", type=int, default=8089)
    parser.add_argument("--winrm-port", type=int, default=5985)
    parser.add_argument("--winrm-transport", default="ntlm", choices=["ntlm", "basic", "kerberos", "credssp"])
    parser.add_argument("--workers", type=int, default=25, help="Parallel worker count")
    parser.add_argument("--json-report", default="", help="Optional output path for JSON summary")
    args = parser.parse_args()

    if args.linux_pass is None:
        args.linux_pass = getpass.getpass("Linux password: ")

    vms = build_vm_rows_from_args(args)
    if not vms:
        raise RuntimeError("No valid VM rows found in XLSX")

    with tempfile.TemporaryDirectory(prefix="trellix_zip_") as tmp:
        zip_name = Path(args.linux_zip_path).name
        local_zip = os.path.join(tmp, zip_name)

        print(f"Downloading ZIP from Linux: {args.linux_user}@{args.linux_host}:{args.linux_zip_path}")
        download_zip_from_linux(args.linux_host, args.linux_user, args.linux_pass, args.linux_zip_path, local_zip)
        print(f"Downloaded to: {local_zip}")

        if args.http_host:
            http_host = args.http_host
        else:
            http_host = pick_local_ip()

        # Validate host format early if IP-like.
        try:
            ipaddress.ip_address(http_host)
        except ValueError:
            pass

        server = start_http_server(tmp, args.http_port)
        package_url = f"http://{http_host}:{args.http_port}/{zip_name}"
        print(f"Temporary package URL: {package_url}")
        print(
            f"Starting parallel WinRM transfer+unzip+install for {len(vms)} VM(s) "
            f"with workers={args.workers}"
        )

        ok = []
        fail = []
        full_results = []
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
                future_map = {
                    ex.submit(
                        transfer_unzip_install_and_check,
                        vm,
                        package_url,
                        zip_name,
                        args.winrm_transport,
                        args.winrm_port,
                    ): vm
                    for vm in vms
                }

                for fut in concurrent.futures.as_completed(future_map):
                    vm = future_map[fut]
                    ip = vm["ip"]
                    try:
                        success, details = fut.result()
                        result_state = details.get("result") or "UNKNOWN"
                        service_name = details.get("service_name") or "N/A"
                        service_status = details.get("service_status") or "N/A"
                        version = details.get("version") or "unknown"
                        full_results.append({"ip": ip, "success": success, **details})

                        if success:
                            ok.append(ip)
                            print(
                                f"[OK] {ip} -> {result_state} | "
                                f"service={service_name}:{service_status} | version={version}"
                            )
                        else:
                            fail.append((ip, details.get("raw") or result_state or "Unknown error"))
                            print(
                                f"[FAIL] {ip} -> {result_state} | "
                                f"service={service_name}:{service_status}"
                            )
                    except Exception as e:
                        fail.append((ip, str(e)))
                        full_results.append({"ip": ip, "success": False, "result": "EXCEPTION", "raw": str(e)})
                        print(f"[FAIL] {ip} -> {e}")
        finally:
            server.shutdown()
            server.server_close()

        print("\nSummary")
        print(f"Success: {len(ok)}")
        print(f"Failed : {len(fail)}")
        if fail:
            print("Failed targets:")
            for ip, msg in fail:
                print(f"- {ip}: {msg}")

        if args.json_report:
            report_dir = os.path.dirname(os.path.abspath(args.json_report))
            if report_dir and not os.path.exists(report_dir):
                os.makedirs(report_dir, exist_ok=True)
            with open(args.json_report, "w", encoding="utf-8") as f:
                json.dump(
                    {
                        "linux_source": {
                            "host": args.linux_host,
                            "user": args.linux_user,
                            "zip_path": args.linux_zip_path,
                        },
                        "package_url": package_url,
                        "success_count": len(ok),
                        "failed_count": len(fail),
                        "results": full_results,
                    },
                    f,
                    indent=2,
                )
            print(f"JSON report written to: {args.json_report}")


if __name__ == "__main__":
    main()
